#!/usr/bin/env python3
"""
公示表 vs 电子报告 交叉比对脚本
比较 Publicsheet/公示表/ 中的数据汇总表与 Publicsheet/电子报告/ 中的独立报告是否一致。
"""

import os
import re
import openpyxl
import xlrd
from collections import defaultdict

BASE_DIR = '/root/projects/report-verification/Publicsheet'
PUBLIC_DIR = os.path.join(BASE_DIR, '公示表')
REPORT_DIR = os.path.join(BASE_DIR, '电子报告')


def normalize_item_name(name):
    """统一检测项目名称，用于匹配"""
    if not name:
        return ''
    name = str(name).strip()
    # 去掉所有括号及其内容（中英文括号），包括单位、计量说明等
    # 先处理嵌套情况：去掉 (xxx) 和 （xxx）
    name = re.sub(r'[（(][^）)]*[）)]', '', name)
    # 再清理一遍，处理可能残留的不匹配括号
    name = re.sub(r'[（(][^）)]*$', '', name)  # 去掉未闭合的左括号到末尾
    name = re.sub(r'^[^（(]*[）)]', '', name)  # 去掉未开的右括号
    # 去掉残留的单独括号字符
    name = name.replace('(', '').replace(')', '').replace('（', '').replace('）', '')
    # 统一名称变体
    name = name.replace('挥发酚类', '挥发酚')
    name = name.replace('阴离子合成洗涤剂', '阴离子表面活性剂')
    name = name.replace('氨氮', '氨')
    name = name.replace('总alpha放射性', '总α放射性')
    name = name.replace('总beta放射性', '总β放射性')
    name = name.strip()
    return name


def normalize_value(val):
    """统一检测值的表示，用于比较"""
    if val is None:
        return None
    val = str(val).strip()
    if val in ('', '/', '-', 'None'):
        return None
    # 统一全角半角 < 符号
    val = val.replace('＜', '<')
    val = val.replace('≤', '<=')
    # 去掉多余空格
    val = re.sub(r'\s+', '', val)
    # 尝试转为数值统一精度
    try:
        f = float(val)
        return f
    except ValueError:
        pass
    # 处理 <数值 格式
    m = re.match(r'^<([\d.]+)$', val)
    if m:
        return f'<{m.group(1)}'
    return val


def values_match(v1, v2):
    """比较两个检测值是否一致"""
    n1 = normalize_value(v1)
    n2 = normalize_value(v2)

    if n1 is None and n2 is None:
        return True
    if n1 is None or n2 is None:
        return False

    # 都是数值
    if isinstance(n1, float) and isinstance(n2, float):
        if n1 == 0 and n2 == 0:
            return True
        if n1 == 0 or n2 == 0:
            return abs(n1 - n2) < 1e-9
        # 容差比较
        return abs(n1 - n2) / max(abs(n1), abs(n2)) < 0.001

    # 都是字符串
    s1 = str(n1).strip()
    s2 = str(n2).strip()

    # 处理 <数值 的精度差异: <0.01 vs <0.010
    m1 = re.match(r'^<([\d.]+)$', s1)
    m2 = re.match(r'^<([\d.]+)$', s2)
    if m1 and m2:
        return abs(float(m1.group(1)) - float(m2.group(1))) < 1e-9

    # 一个是数值一个是 <数值
    if isinstance(n1, float) and m2:
        return False
    if isinstance(n2, float) and m1:
        return False

    # 文本比较
    text_equiv = {
        '无': '无',
        '无异臭、异味': '无异臭、异味',
        '无异臭异味': '无异臭、异味',
        '未检出': '未检出',
        '0': '未检出',
    }
    s1_norm = text_equiv.get(s1, s1)
    s2_norm = text_equiv.get(s2, s2)
    return s1_norm == s2_norm


# ===================== 读取公示表 =====================

def read_public_sheets():
    """
    读取所有公示表，返回:
    {sample_id: {item_name: value, ...}, ...}
    同时记录 sample_id -> (公司名, 水质类型, 采样地点)
    """
    data = {}
    sample_info = {}

    for fname in sorted(os.listdir(PUBLIC_DIR)):
        if not fname.endswith('.xlsx'):
            continue
        fpath = os.path.join(PUBLIC_DIR, fname)
        company = fname.replace('数据汇总表.xlsx', '')
        wb = openpyxl.load_workbook(fpath, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 判断水质类型
            if '出' in sheet_name:
                water_type = '出厂水'
            elif '原' in sheet_name:
                water_type = '原水'
            elif '管' in sheet_name:
                water_type = '管网水'
            else:
                water_type = sheet_name

            max_row = ws.max_row
            max_col = ws.max_column

            # 行2: 样品编号行
            sample_ids = []
            for col in range(2, max_col + 1):
                sid = ws.cell(row=2, column=col).value
                if sid:
                    sid = str(sid).strip()
                sample_ids.append(sid)

            # 行3: 采样地点行
            locations = []
            for col in range(2, max_col + 1):
                loc = ws.cell(row=3, column=col).value
                if loc:
                    loc = str(loc).strip()
                locations.append(loc)

            # 行4起: 检测项目数据
            for row in range(4, max_row + 1):
                item_name_raw = ws.cell(row=row, column=1).value
                if not item_name_raw:
                    continue
                item_name_raw = str(item_name_raw).strip()
                if '以下无检测数据' in item_name_raw:
                    break
                item_key = normalize_item_name(item_name_raw)
                if not item_key:
                    continue

                for col_idx in range(len(sample_ids)):
                    sid = sample_ids[col_idx]
                    if not sid:
                        continue
                    val = ws.cell(row=row, column=col_idx + 2).value
                    if val is not None:
                        val = str(val).strip()

                    key = (sid, water_type)
                    if key not in data:
                        data[key] = {}
                    data[key][item_key] = val
                    loc = locations[col_idx] if col_idx < len(locations) else ''
                    sample_info[key] = (company, water_type, loc)

        wb.close()

    return data, sample_info


# ===================== 读取电子报告 =====================

def read_report_xlsx(fpath):
    """读取xlsx格式的电子报告，返回 (sample_id, water_type, location, {item: value})"""
    wb = openpyxl.load_workbook(fpath, data_only=True)
    sheets = wb.sheetnames

    sample_id = None
    water_type = None
    location = None
    items = {}

    # Sheet 2: 样品信息
    if len(sheets) >= 2:
        ws2 = wb[sheets[1]]
        for row in range(1, min(20, ws2.max_row + 1)):
            for col in range(1, min(10, ws2.max_column + 1)):
                cell_val = ws2.cell(row=row, column=col).value
                if cell_val:
                    cell_str = str(cell_val).strip()
                    if '样品编号' in cell_str:
                        # 样品编号通常在同行的下一列或下两列
                        for c2 in range(col + 1, min(col + 4, ws2.max_column + 1)):
                            v = ws2.cell(row=row, column=c2).value
                            if v and str(v).strip().startswith('W'):
                                sample_id = str(v).strip()
                                break
                    if '样品类型' in cell_str:
                        for c2 in range(col + 1, min(col + 4, ws2.max_column + 1)):
                            v = ws2.cell(row=row, column=c2).value
                            if v:
                                water_type = str(v).strip()
                                break
                    if '采样地点' in cell_str:
                        for c2 in range(col + 1, min(col + 4, ws2.max_column + 1)):
                            v = ws2.cell(row=row, column=c2).value
                            if v:
                                location = str(v).strip()
                                break

    # Sheet 3 & 4: 检测结果
    for si in range(2, min(4, len(sheets))):
        ws = wb[sheets[si]]
        for row in range(1, ws.max_row + 1):
            seq = ws.cell(row=row, column=1).value
            # 检查序号列是否为数字
            try:
                seq_num = int(float(str(seq)))
            except (ValueError, TypeError):
                continue
            if seq_num < 1 or seq_num > 100:
                continue

            item_name_raw = ws.cell(row=row, column=2).value
            if not item_name_raw:
                continue
            item_name_raw = str(item_name_raw).strip()
            item_key = normalize_item_name(item_name_raw)
            if not item_key:
                continue

            result_val = ws.cell(row=row, column=4).value
            if result_val is not None:
                result_val = str(result_val).strip()

            items[item_key] = result_val

    wb.close()
    return sample_id, water_type, location, items


def read_report_xls(fpath):
    """读取xls格式的电子报告"""
    wb = xlrd.open_workbook(fpath)
    sheets = wb.sheet_names()

    sample_id = None
    water_type = None
    location = None
    items = {}

    # Sheet 2: 样品信息
    if len(sheets) >= 2:
        ws2 = wb.sheet_by_index(1)
        for row in range(min(20, ws2.nrows)):
            for col in range(min(10, ws2.ncols)):
                cell_val = ws2.cell_value(row, col)
                if cell_val:
                    cell_str = str(cell_val).strip()
                    if '样品编号' in cell_str:
                        for c2 in range(col + 1, min(col + 4, ws2.ncols)):
                            v = ws2.cell_value(row, c2)
                            if v and str(v).strip().startswith('W'):
                                sample_id = str(v).strip()
                                break
                    if '样品类型' in cell_str:
                        for c2 in range(col + 1, min(col + 4, ws2.ncols)):
                            v = ws2.cell_value(row, c2)
                            if v:
                                water_type = str(v).strip()
                                break
                    if '采样地点' in cell_str:
                        for c2 in range(col + 1, min(col + 4, ws2.ncols)):
                            v = ws2.cell_value(row, c2)
                            if v:
                                location = str(v).strip()
                                break

    # Sheet 3 & 4: 检测结果
    for si in range(2, min(4, len(sheets))):
        ws = wb.sheet_by_index(si)
        for row in range(ws.nrows):
            seq = ws.cell_value(row, 0)
            try:
                seq_num = int(float(str(seq)))
            except (ValueError, TypeError):
                continue
            if seq_num < 1 or seq_num > 100:
                continue

            item_name_raw = ws.cell_value(row, 1)
            if not item_name_raw:
                continue
            item_name_raw = str(item_name_raw).strip()
            item_key = normalize_item_name(item_name_raw)
            if not item_key:
                continue

            result_val = ws.cell_value(row, 3)
            if result_val is not None and result_val != '':
                result_val = str(result_val).strip()
            else:
                result_val = None

            items[item_key] = result_val

    return sample_id, water_type, location, items


def classify_water_type(wtype):
    """归一化水质类型"""
    if not wtype:
        return ''
    if '出' in wtype:
        return '出厂水'
    if '原' in wtype:
        return '原水'
    if '管' in wtype:
        return '管网水'
    return wtype


def read_all_reports():
    """读取所有电子报告，返回 {(sample_id, water_type): {item: value}} 和元信息"""
    report_data = {}
    report_info = {}
    errors = []

    for fname in sorted(os.listdir(REPORT_DIR)):
        if not fname.endswith(('.xlsx', '.xls')):
            continue
        fpath = os.path.join(REPORT_DIR, fname)
        try:
            if fname.endswith('.xlsx'):
                sid, wtype, loc, items = read_report_xlsx(fpath)
            else:
                sid, wtype, loc, items = read_report_xls(fpath)

            if not sid:
                errors.append(f"[无法提取样品编号] {fname}")
                continue
            if not items:
                errors.append(f"[无检测数据] {fname} (样品编号: {sid})")
                continue

            wtype_norm = classify_water_type(wtype)
            key = (sid, wtype_norm)

            if key in report_data:
                errors.append(f"[样品编号+类型重复] {sid}/{wtype_norm} 在 {fname} 中重复出现")

            report_data[key] = items
            report_info[key] = (fname, wtype, loc)

        except Exception as e:
            errors.append(f"[读取失败] {fname}: {e}")

    return report_data, report_info, errors


# ===================== 交叉比对 =====================

def cross_verify():
    print("=" * 80)
    print("公示表 vs 电子报告 交叉比对")
    print("=" * 80)

    print("\n[1] 读取公示表...")
    pub_data, pub_info = read_public_sheets()
    print(f"    公示表中共有 {len(pub_data)} 个样品")

    print("\n[2] 读取电子报告...")
    rpt_data, rpt_info, rpt_errors = read_all_reports()
    print(f"    电子报告中共有 {len(rpt_data)} 个样品")

    if rpt_errors:
        print(f"\n    读取报告时的问题 ({len(rpt_errors)} 项):")
        for e in rpt_errors:
            print(f"      - {e}")

    # 匹配检查
    pub_ids = set(pub_data.keys())
    rpt_ids = set(rpt_data.keys())

    matched = pub_ids & rpt_ids
    pub_only = pub_ids - rpt_ids
    rpt_only = rpt_ids - pub_ids

    print(f"\n[3] 样品匹配情况:")
    print(f"    匹配成功: {len(matched)} 个")
    print(f"    仅在公示表中: {len(pub_only)} 个")
    print(f"    仅在电子报告中: {len(rpt_only)} 个")

    if pub_only:
        print(f"\n    公示表中有但电子报告中没有的样品:")
        for key in sorted(pub_only):
            info = pub_info.get(key, ('', '', ''))
            print(f"      {key[0]} / {key[1]} ({info[0]}, {info[2]})")

    if rpt_only:
        print(f"\n    电子报告中有但公示表中没有的样品:")
        for key in sorted(rpt_only):
            info = rpt_info.get(key, ('', '', ''))
            print(f"      {key[0]} / {key[1]} ({info[0]}, {info[2]})")

    # 逐样品逐项比对
    print(f"\n[4] 逐项数据比对 ({len(matched)} 个匹配样品):")
    print("-" * 80)

    total_items_checked = 0
    total_mismatches = 0
    total_missing_in_pub = 0
    total_missing_in_rpt = 0
    mismatch_details = []

    for key in sorted(matched):
        pub_items = pub_data[key]
        rpt_items = rpt_data[key]
        pub_i = pub_info.get(key, ('', '', ''))
        rpt_i = rpt_info.get(key, ('', '', ''))
        sid = f"{key[0]} / {key[1]}"

        all_keys = set(pub_items.keys()) | set(rpt_items.keys())
        sample_mismatches = []

        for key in sorted(all_keys):
            pv = pub_items.get(key)
            rv = rpt_items.get(key)

            # 跳过两边都空的
            if normalize_value(pv) is None and normalize_value(rv) is None:
                continue

            total_items_checked += 1

            if normalize_value(pv) is None:
                total_missing_in_pub += 1
                sample_mismatches.append((key, '[公示表缺失]', rv))
                continue
            if normalize_value(rv) is None:
                total_missing_in_rpt += 1
                sample_mismatches.append((key, pv, '[报告缺失]'))
                continue

            if not values_match(pv, rv):
                total_mismatches += 1
                sample_mismatches.append((key, pv, rv))

        if sample_mismatches:
            mismatch_details.append((sid, pub_i, rpt_i, sample_mismatches))

    # 输出结果
    if not mismatch_details:
        print("\n  所有匹配样品的检测数据完全一致！")
    else:
        print(f"\n  发现数据不一致的样品: {len(mismatch_details)} 个")
        print(f"  总检查项数: {total_items_checked}")
        print(f"  数值不一致: {total_mismatches} 项")
        print(f"  公示表缺失: {total_missing_in_pub} 项")
        print(f"  报告缺失:   {total_missing_in_rpt} 项")

        print("\n" + "=" * 80)
        print("详细不一致清单")
        print("=" * 80)

        for sid, pub_i, rpt_i, mismatches in mismatch_details:
            print(f"\n样品编号: {sid}")
            print(f"  公示表: {pub_i[0]} / {pub_i[1]} / {pub_i[2]}")
            print(f"  报告:   {rpt_i[0]} / {rpt_i[1]} / {rpt_i[2]}")
            print(f"  不一致项数: {len(mismatches)}")
            print(f"  {'检测项目':<25s} {'公示表值':<20s} {'报告值':<20s}")
            print(f"  {'-'*25} {'-'*20} {'-'*20}")
            for item, pv, rv in mismatches:
                print(f"  {item:<25s} {str(pv):<20s} {str(rv):<20s}")

    # 汇总
    print("\n" + "=" * 80)
    print("汇总")
    print("=" * 80)
    print(f"公示表样品数:       {len(pub_data)}")
    print(f"电子报告样品数:     {len(rpt_data)}")
    print(f"匹配样品数:         {len(matched)}")
    print(f"仅公示表:           {len(pub_only)}")
    print(f"仅电子报告:         {len(rpt_only)}")
    print(f"总检查项数:         {total_items_checked}")
    print(f"数值不一致:         {total_mismatches}")
    print(f"公示表缺失:         {total_missing_in_pub}")
    print(f"报告缺失:           {total_missing_in_rpt}")


if __name__ == '__main__':
    cross_verify()
