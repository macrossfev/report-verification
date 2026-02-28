#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
水质检测报告文件全面分析脚本
扫描所有 .xlsx / .xls 报告，按类别输出待确认问题清单。
"""

import argparse
import os, re, sys, traceback
from collections import defaultdict, Counter
from datetime import datetime

import openpyxl
import xlrd

# ──────────────────────────── helpers ────────────────────────────

def extract_number_prefix(fname):
    """Return the leading numeric string of a filename, e.g. '0001'."""
    m = re.match(r'^(\d+)', fname)
    return m.group(1) if m else None

def classify_water_type(fname):
    """Classify report type from filename."""
    if '二次供水' in fname:
        return '二次供水'
    if '农饮水' in fname or '生活饮用水' in fname:
        return '农饮水'
    if '转供水' in fname:
        return '转供水'
    if '日检九项' in fname:
        return '日检九项'
    if '送检' in fname:
        return '送检'
    if '高锰酸盐指数' in fname:
        return '高锰酸盐指数'
    if '原水' in fname:
        return '原水'
    if '出厂水' in fname:
        return '出厂水'
    if '管网' in fname:
        return '管网水'
    return '未知'

def extract_plant_name(fname):
    """Try to extract water plant name from filename."""
    # Remove prefix number
    name = re.sub(r'^\d+', '', fname)
    # Remove extension
    name = re.sub(r'\.(xlsx?|xls)$', '', name)
    # Remove date suffixes like 01.05
    name = re.sub(r'\s*\d{2}\.\d{2}\s*$', '', name)
    # Remove trailing markers
    for tag in ['-送检', '送检', '-荣昌', '荣昌', '日检九项', '高锰酸盐指数',
                '-地表三类', '地表三类', '应急水样', '-应急水样']:
        name = name.replace(tag, '')
    name = name.strip(' -')
    # Try to get the plant name before the water-type bracket
    # e.g. 北门水厂（出厂水） -> 北门水厂
    m = re.match(r'^(.+?水厂|.+?水库|.+?泵站).*', name)
    if m:
        plant = m.group(1)
        # Normalize: remove 管网水 prefix patterns
        plant = re.sub(r'管网水$', '水厂', plant)
        return plant.strip()
    # For things like 小北海（出厂水）
    m = re.match(r'^([^（(]+)', name)
    if m:
        return m.group(1).strip()
    return name.strip()

# ──────────────────── reading Excel data ────────────────────

def read_xlsx_report_info(filepath):
    """Read key metadata from an xlsx file."""
    info = {}
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        info['sheet_count'] = len(wb.sheetnames)
        info['sheet_names'] = wb.sheetnames

        # Page 1 (cover page) - try first sheet
        ws1 = wb[wb.sheetnames[0]]
        info['total_rows_sheet1'] = ws1.max_row
        info['total_cols_sheet1'] = ws1.max_column

        # Extract report number from B1
        b1 = ws1.cell(1, 2).value
        if b1:
            info['report_number_raw'] = str(b1).strip()
            m = re.search(r'第\s*\(\s*(\d+)\s*\)\s*号', str(b1))
            if m:
                info['report_number'] = m.group(1).strip()

        # Extract page info from B2
        b2 = ws1.cell(2, 2).value
        if b2:
            m = re.search(r'共\s*(\d+)\s*页', str(b2))
            if m:
                info['total_pages'] = int(m.group(1))

        # Sample name from C8 or C9 (row 8)
        for r in range(7, 13):
            cv = ws1.cell(r, 3).value
            if cv and ('水' in str(cv) or '【' in str(cv)):
                info['sample_name'] = str(cv).strip()
                break

        # Company from C9
        for r in range(8, 13):
            cv = ws1.cell(r, 3).value
            if cv and '公司' in str(cv):
                info['company'] = str(cv).strip()
                break

        # Report date from C12 or C11
        for r in range(10, 14):
            bv = ws1.cell(r, 2).value
            cv = ws1.cell(r, 3).value
            if bv and '报告编制日期' in str(bv) and cv:
                info['report_date'] = str(cv).strip()
                break

        # Page 2 (检测结果) - try second sheet
        if len(wb.sheetnames) >= 2:
            ws2 = wb[wb.sheetnames[1]]
            info['total_rows_sheet2'] = ws2.max_row

            # Sample type from C3
            c3 = ws2.cell(3, 3).value
            if c3:
                info['sample_type'] = str(c3).strip()

            # Sampler from C4
            c4 = ws2.cell(4, 3).value
            if c4:
                info['sampler'] = str(c4).strip()

            # Sampling date from E4
            e4 = ws2.cell(4, 5).value
            if e4:
                info['sampling_date'] = str(e4).strip()

            # Receipt date from E5
            e5 = ws2.cell(5, 5).value
            if e5:
                info['receipt_date'] = str(e5).strip()

            # Sampling location from C6
            c6 = ws2.cell(6, 3).value
            if c6:
                info['sampling_location'] = str(c6).strip()

            # Sample ID from C8
            c8 = ws2.cell(8, 3).value
            if c8:
                info['sample_id'] = str(c8).strip()

            # Testing date from E8
            e8 = ws2.cell(8, 5).value
            if e8:
                info['testing_date'] = str(e8).strip()

            # Product standard from C9
            c9 = ws2.cell(9, 3).value
            if c9:
                info['product_standard'] = str(c9).strip()

            # Number of test items from C10
            c10 = ws2.cell(10, 3).value
            if c10:
                info['test_items_desc'] = str(c10).strip()
                m = re.search(r'(\d+)\s*项', str(c10))
                if m:
                    info['test_item_count'] = int(m.group(1))

            # Conclusion from B13
            b13 = ws2.cell(13, 2).value
            if b13:
                info['conclusion'] = str(b13).strip()

        # Page 3+ (检测数据) - collect test items
        test_items = []
        for si in range(2, len(wb.sheetnames)):
            ws = wb[wb.sheetnames[si]]
            for r in range(1, ws.max_row + 1):
                a_val = ws.cell(r, 1).value
                b_val = ws.cell(r, 2).value
                d_val = ws.cell(r, 4).value
                if a_val is not None and b_val is not None:
                    try:
                        seq = int(float(str(a_val)))
                        if 1 <= seq <= 100 and b_val:
                            item = {
                                'seq': seq,
                                'name': str(b_val).strip(),
                                'unit': str(ws.cell(r, 3).value or '').strip(),
                                'result': format_cell_number(d_val, ws.cell(r, 4).number_format) if d_val is not None else '',
                                'standard': str(ws.cell(r, 5).value or '').strip(),
                                'method': str(ws.cell(r, 6).value or '').strip(),
                            }
                            test_items.append(item)
                    except (ValueError, TypeError):
                        pass
        info['test_items'] = test_items
        wb.close()
    except Exception as e:
        info['read_error'] = f"{type(e).__name__}: {e}"
    return info


def read_xls_report_info(filepath):
    """Read key metadata from an xls file."""
    info = {}
    try:
        wb = xlrd.open_workbook(filepath, formatting_info=True)
        info['sheet_count'] = wb.nsheets
        info['sheet_names'] = wb.sheet_names()

        ws1 = wb.sheet_by_index(0)
        info['total_rows_sheet1'] = ws1.nrows
        info['total_cols_sheet1'] = ws1.ncols

        # Report number from B1 (row 0, col 1)
        if ws1.nrows > 0 and ws1.ncols > 1:
            b1 = ws1.cell_value(0, 1)
            if b1:
                info['report_number_raw'] = str(b1).strip()
                m = re.search(r'第\s*\(\s*(\d+)\s*\)\s*号', str(b1))
                if m:
                    info['report_number'] = m.group(1).strip()

        # Page info from B2 (row 1, col 1)
        if ws1.nrows > 1 and ws1.ncols > 1:
            b2 = ws1.cell_value(1, 1)
            if b2:
                m = re.search(r'共\s*(\d+)\s*页', str(b2))
                if m:
                    info['total_pages'] = int(m.group(1))

        # Sample name from C8 (row 7, col 2)
        for r in range(6, min(12, ws1.nrows)):
            if ws1.ncols > 2:
                cv = ws1.cell_value(r, 2)
                if cv and ('水' in str(cv) or '【' in str(cv)):
                    info['sample_name'] = str(cv).strip()
                    break

        # Company
        for r in range(7, min(12, ws1.nrows)):
            if ws1.ncols > 2:
                cv = ws1.cell_value(r, 2)
                if cv and '公司' in str(cv):
                    info['company'] = str(cv).strip()
                    break

        # Report date
        for r in range(9, min(14, ws1.nrows)):
            if ws1.ncols > 1:
                bv = ws1.cell_value(r, 1) if ws1.ncols > 1 else ''
                cv = ws1.cell_value(r, 2) if ws1.ncols > 2 else ''
                if bv and '报告编制日期' in str(bv) and cv:
                    info['report_date'] = str(cv).strip()
                    break

        # Page 2
        if wb.nsheets >= 2:
            ws2 = wb.sheet_by_index(1)
            info['total_rows_sheet2'] = ws2.nrows

            def sv(r, c):
                if r < ws2.nrows and c < ws2.ncols:
                    return ws2.cell_value(r, c)
                return None

            c3 = sv(2, 2)
            if c3:
                info['sample_type'] = str(c3).strip()

            c4 = sv(3, 2)
            if c4:
                info['sampler'] = str(c4).strip()

            e4 = sv(3, 4)
            if e4:
                info['sampling_date'] = str(e4).strip()

            e5 = sv(4, 4)
            if e5:
                info['receipt_date'] = str(e5).strip()

            c6 = sv(5, 2)
            if c6:
                info['sampling_location'] = str(c6).strip()

            c8 = sv(7, 2)
            if c8:
                info['sample_id'] = str(c8).strip()

            e8 = sv(7, 4)
            if e8:
                info['testing_date'] = str(e8).strip()

            c9 = sv(8, 2)
            if c9:
                info['product_standard'] = str(c9).strip()

            c10 = sv(9, 2)
            if c10:
                info['test_items_desc'] = str(c10).strip()
                m = re.search(r'(\d+)\s*项', str(c10))
                if m:
                    info['test_item_count'] = int(m.group(1))

            b13 = sv(12, 1)
            if b13:
                info['conclusion'] = str(b13).strip()

        # Test items from page 3+
        test_items = []
        for si in range(2, wb.nsheets):
            ws = wb.sheet_by_index(si)
            for r in range(ws.nrows):
                if ws.ncols >= 6:
                    a_val = ws.cell_value(r, 0)
                    b_val = ws.cell_value(r, 1)
                    d_val = ws.cell_value(r, 3)
                    if a_val not in ('', None) and b_val not in ('', None):
                        try:
                            seq = int(float(str(a_val)))
                            if 1 <= seq <= 100 and b_val:
                                # Get xls cell number format for result column
                                d_fmt = None
                                try:
                                    xf_idx = ws.cell_xf_index(r, 3)
                                    fmt_key = wb.xf_list[xf_idx].format_key
                                    d_fmt = wb.format_map.get(fmt_key, None)
                                    if d_fmt:
                                        d_fmt = d_fmt.format_str
                                except Exception:
                                    pass
                                item = {
                                    'seq': seq,
                                    'name': str(b_val).strip(),
                                    'unit': str(ws.cell_value(r, 2) or '').strip(),
                                    'result': format_cell_number(d_val, d_fmt) if d_val not in ('', None) else '',
                                    'standard': str(ws.cell_value(r, 4) or '').strip(),
                                    'method': str(ws.cell_value(r, 5) or '').strip(),
                                }
                                test_items.append(item)
                        except (ValueError, TypeError):
                            pass
        info['test_items'] = test_items

    except Exception as e:
        info['read_error'] = f"{type(e).__name__}: {e}"
    return info


# ──────────────── Original Record Handling ────────────────

def find_original_record_file(directory):
    """Find original record file (e.g., 260205-1-25.xlsx) in directory."""
    for f in sorted(os.listdir(directory)):
        if re.match(r'^\d{6}-\d+-\d+\.xlsx$', f):
            return os.path.join(directory, f)
    return None


def clean_item_name(raw):
    """Clean test item name for matching."""
    s = str(raw).strip()
    s = re.sub(r'\s*\n\s*', '', s)
    # Remove trailing punctuation
    s = re.sub(r'[、，,]+$', '', s).strip()
    # Remove unit suffixes like (mg/L)
    s = re.sub(r'\s*[\(（][^)）]*[\)）]\s*$', '', s).strip()
    s = re.sub(r'\s*[\(（][^)）]*$', '', s).strip()
    # Remove spaces between CJK characters (e.g., 氰 化 物 -> 氰化物)
    s = re.sub(r'(?<=[\u4e00-\u9fff])\s+(?=[\u4e00-\u9fff])', '', s)
    s = re.sub(r'\s{2,}', ' ', s).strip()
    return s


def format_cell_number(value, number_format=None):
    """Format numeric cell preserving Excel decimal places (e.g., 0.30 stays 0.30)."""
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    if not isinstance(value, (int, float)):
        return str(value).strip()
    if number_format and number_format not in ('General', '@', '', '0'):
        if '.' in number_format:
            after_dot = number_format.split('.')[-1]
            m = re.match(r'[0#?]+', after_dot)
            if m:
                return f"{value:.{len(m.group())}f}"
    if isinstance(value, float) and value == int(value) and abs(value) < 1e10:
        return str(int(value))
    return str(value)


def get_param_value(items, *names):
    """Get numeric value from items dict by fuzzy name matching."""
    for name in names:
        for key, val in items.items():
            if name == key or ((name in key or key in name) and not is_false_substring_match(name, key)):
                s = str(val).replace('＜', '<').strip()
                if s.startswith('<'):
                    return None, key, s
                try:
                    return float(s), key, s
                except (ValueError, TypeError):
                    pass
    return None, None, None


def check_data_logic(items, label):
    """Check logical consistency of test data values."""
    issues = []

    # 1. TDS vs conductivity: ratio should be ~0.4-0.8
    tds, _, tds_r = get_param_value(items, '溶解性总固体')
    ec, _, ec_r = get_param_value(items, '电导率')
    if tds and ec and ec > 0:
        ratio = tds / ec
        if ratio < 0.3 or ratio > 1.0:
            issues.append(f"{label} 溶解性总固体({tds_r})/电导率({ec_r})比值={ratio:.2f}，通常应在0.4-0.8之间")

    # 2. Total Cr >= Cr(VI)
    cr_t, _, cr_r = get_param_value(items, '总铬')
    cr6, _, cr6_r = get_param_value(items, '铬(六价)', '六价铬')
    if cr_t is not None and cr6 is not None and cr6 > cr_t:
        issues.append(f"{label} 铬(六价)({cr6_r})大于总铬({cr_r})，逻辑矛盾")

    # 3. High Fe/Mn -> color should not be below detection limit
    fe, _, _ = get_param_value(items, '铁')
    mn, _, _ = get_param_value(items, '锰')
    color_raw = None
    for k, v in items.items():
        if '色度' in k:
            color_raw = str(v).strip()
            break
    if color_raw and (color_raw.startswith('<') or color_raw.startswith('＜')):
        high_parts = []
        if fe and fe > 0.3:
            high_parts.append(f"铁={fe}")
        if mn and mn > 0.1:
            high_parts.append(f"锰={mn}")
        if high_parts:
            issues.append(f"{label} {'、'.join(high_parts)}偏高但色度低于检出限({color_raw})，需确认")

    # 4. Total N >= NH3-N + NO3-N + NO2-N
    tn, _, tn_r = get_param_value(items, '总氮')
    nh3, _, nh3_r = get_param_value(items, '氨氮', '氨(以N计)', '氨')
    no3, _, no3_r = get_param_value(items, '硝酸盐')
    no2, _, no2_r = get_param_value(items, '亚硝酸盐')
    if tn is not None:
        comp = sum(v for v in [nh3 or 0, no3 or 0, no2 or 0])
        if comp > tn * 1.1 and comp > 0:
            issues.append(
                f"{label} 氨氮({nh3_r})+硝酸盐氮({no3_r})+亚硝酸盐氮({no2_r})"
                f"之和({comp:.3f})大于总氮({tn_r})，逻辑矛盾")

    # 5. High DO -> NH3 should be low
    do, _, do_r = get_param_value(items, '溶解氧')
    if do and do > 7 and nh3 and nh3 > 0.5:
        issues.append(f"{label} 溶解氧({do_r})较高但氨氮({nh3_r})偏高，需确认")

    # 6. NO2 usually < NH3 and < NO3
    if no2 and no2 > 0:
        if nh3 and nh3 > 0 and no2 > nh3 * 2:
            issues.append(f"{label} 亚硝酸盐氮({no2_r})显著高于氨氮({nh3_r})，异常")
        if no3 and no3 > 0 and no2 > no3 * 2:
            issues.append(f"{label} 亚硝酸盐氮({no2_r})显著高于硝酸盐氮({no3_r})，异常")

    return issues


NAME_ALIAS = {
    '高锰酸盐指数': '高锰酸盐指数(以O2计)',
    '溶解氧': '溶解氧', '化学需氧量': '化学需氧量(COD)',
    '五日生化需氧量': '五日生化需氧量(BOD5)',
    '挥发酚': '挥发酚类(以苯酚计)',
    '六 价 铬': '铬(六价)', '六价铬': '铬(六价)',
    '总硬度': '总硬度(以CaCO3计)',
    '氨': '氨(以N计)', '氨(以N计)': '氨氮(NH3-N)', '硝酸盐': '硝酸盐(以N计)',
    '总α': '总α放射性', '总a': '总α放射性', '总β': '总β放射性',
    '总磷': '总磷(以P计)', '总氮': '总氮(以N计)', '氨氮': '氨氮(NH3-N)',
    '阴离子表面活性剂': '阴离子合成洗涤剂',
}

# Parameter names that should NOT match each other via substring matching
# (short_name, long_name_keyword) — if one name IS the short form and the other CONTAINS the long keyword, reject
CONFUSABLE_PARAMS = [
    ('锰', '高锰酸盐'),
]


def is_false_substring_match(name_a, name_b):
    """Return True if name_a and name_b should NOT be considered a substring match."""
    clean_a = re.sub(r'[\(（].*?[\)）]', '', name_a).strip()
    clean_b = re.sub(r'[\(（].*?[\)）]', '', name_b).strip()
    for short, long_kw in CONFUSABLE_PARAMS:
        if clean_a == short and long_kw in clean_b:
            return True
        if clean_b == short and long_kw in clean_a:
            return True
    return False


def normalize_method(method_str):
    """Normalize detection method string for comparison (ignore formatting differences)."""
    s = str(method_str).strip()
    s = s.replace('\n', ' ').replace('\r', '')
    s = s.replace('\u3000', ' ')  # full-width space
    s = re.sub(r'\s+', ' ', s)
    # Normalize full-width punctuation to half-width
    s = s.replace('（', '(').replace('）', ')').replace('，', ',')
    s = s.replace('：', ':').replace('；', ';').replace('、', ',')
    # Remove spaces between ASCII and CJK characters
    s = re.sub(r'(?<=[\x21-\x7e])\s+(?=[\u4e00-\u9fff])', '', s)
    s = re.sub(r'(?<=[\u4e00-\u9fff])\s+(?=[\x21-\x7e])', '', s)
    return s.strip()


def count_digits(value_str):
    """Count total digit count in a number string (all digits including zeros).

    '17.6' -> 3, '7.63' -> 3, '0.64' -> 3, '1.00' -> 3, '0.005' -> 4, '100' -> 3
    """
    s = value_str.strip()
    if s.startswith('-') or s.startswith('+'):
        s = s[1:]
    return len(s.replace('.', ''))


def read_original_record(filepath):
    """Read original record: sample registry (Sheet1) + all test data."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sid_pattern = r'[WKM]\d{6}C\d+'

    # ── Sheet1: Sample Registry ──
    registry = []
    ws1 = wb[wb.sheetnames[0]]

    # Auto-detect header: find row with "样品编号" and determine column layout
    sid_col = None
    company_col = None
    desc_col = None
    samp_code_col = None
    data_start_row = None

    for r in range(1, min(ws1.max_row + 1, 8)):
        for c in range(1, ws1.max_column + 1):
            v = ws1.cell(r, c).value
            if v and '样品编号' in str(v):
                sid_col = c
                data_start_row = r + 1
            elif v and '被检单位' in str(v):
                company_col = c
            elif v and '采样地点' in str(v):
                desc_col = c
            elif v and '采样编号' in str(v):
                samp_code_col = c

    # Fallback defaults
    if sid_col is None:
        sid_col = 5
    if company_col is None:
        company_col = 2
    if desc_col is None:
        desc_col = 3
    if samp_code_col is None:
        samp_code_col = 4
    if data_start_row is None:
        data_start_row = 4

    for r in range(data_start_row, ws1.max_row + 1):
        sample_id = ws1.cell(r, sid_col).value
        if sample_id and re.match(sid_pattern, str(sample_id).strip()):
            company = ws1.cell(r, company_col).value
            # 如果当前行无 company，向上查找（合并单元格场景）
            if not company:
                for rr in range(r - 1, data_start_row - 1, -1):
                    cv = ws1.cell(rr, company_col).value
                    if cv:
                        company = cv
                        break
            registry.append({
                'seq': ws1.cell(r, 1).value,
                'company': str(company).strip() if company else '',
                'description': str(ws1.cell(r, desc_col).value or '').strip(),
                'sampling_code': str(ws1.cell(r, samp_code_col).value or '').strip(),
                'sample_id': str(sample_id).strip(),
            })

    # ── Test Data from remaining sheets ──
    test_data = defaultdict(dict)

    for si in range(1, len(wb.sheetnames)):
        ws = wb[wb.sheetnames[si]]
        if ws.max_row < 3:
            continue

        # Detect layout: samples in columns (A) or samples in rows (B)
        sample_cols = {}
        header_row = 0
        for hr in [2, 3]:
            for c in range(1, ws.max_column + 1):
                v = ws.cell(hr, c).value
                if v and re.match(sid_pattern, str(v).strip()):
                    sample_cols[c] = str(v).strip()
                    header_row = max(header_row, hr)

        if sample_cols:
            # Layout A: item names in col 1, sample values in columns
            for r in range(header_row + 1, ws.max_row + 1):
                item_cell = ws.cell(r, 1).value
                if not item_cell:
                    continue
                cname = clean_item_name(item_cell)
                if not cname or '项' in cname and '目' in cname:
                    continue
                for c, sid in sample_cols.items():
                    cell = ws.cell(r, c)
                    val = cell.value
                    if val is not None and str(val).strip():
                        formatted = format_cell_number(val, cell.number_format)
                        test_data[sid][cname] = re.sub(r'[、，,]+$', '', formatted)
        else:
            # Layout B: sample IDs in col 1, item names in header columns
            item_cols = {}
            for hr in [2, 3]:
                for c in range(2, ws.max_column + 1):
                    v = ws.cell(hr, c).value
                    if v:
                        cname = clean_item_name(v)
                        if cname and len(cname) > 1:
                            item_cols[c] = cname
            if item_cols:
                for r in range(4, ws.max_row + 1):
                    v = ws.cell(r, 1).value
                    if v and re.match(sid_pattern, str(v).strip()):
                        sid = str(v).strip()
                        for c, iname in item_cols.items():
                            cell = ws.cell(r, c)
                            val = cell.value
                            if val is not None and str(val).strip():
                                formatted = format_cell_number(val, cell.number_format)
                                test_data[sid][iname] = re.sub(r'[、，,]+$', '', formatted)

    wb.close()
    return registry, dict(test_data)


def classify_sample_water_type(desc):
    """Classify water type from sample description."""
    if '二次供水' in desc:
        return '二次供水'
    if '农饮水' in desc or '农村' in desc:
        return '农饮水'
    if '管网末梢' in desc:
        return '管网末梢水'
    if '管网' in desc:
        return '管网水'
    if '出厂水' in desc:
        return '出厂水'
    if '原水' in desc:
        return '原水'
    return '未知'


def extract_plant_from_desc(desc):
    """Extract plant name from sample description."""
    s = desc
    for tag in ['出厂水', '管网末梢水', '管网水', '原水', '农饮水', '二次供水']:
        s = s.replace(tag, '')
    s = re.sub(r'[\(（][^)）]*[\)）]', '', s)
    return s.strip()


def find_matching_report_item(test_items, orig_name):
    """Find matching test item in report by name."""
    for item in test_items:
        if item['name'] == orig_name:
            return item
    alias = NAME_ALIAS.get(orig_name)
    if alias:
        for item in test_items:
            if item['name'] == alias:
                return item
    for item in test_items:
        if (orig_name in item['name'] or item['name'] in orig_name) and not is_false_substring_match(orig_name, item['name']):
            return item
        clean_o = re.sub(r'[\(（].*?[\)）]', '', orig_name).strip()
        clean_i = re.sub(r'[\(（].*?[\)）]', '', item['name']).strip()
        if clean_o and clean_i and len(clean_o) > 1 and (clean_o in clean_i or clean_i in clean_o) and not is_false_substring_match(orig_name, item['name']):
            return item
    return None


def vals_match(orig_val, report_val):
    """Compare original record value with report value — strict exact match."""
    if orig_val is None or report_val is None:
        return True
    o = re.sub(r'[、，,]+$', '', str(orig_val).strip()).replace('＜', '<')
    r = re.sub(r'[、，,]+$', '', str(report_val).strip()).replace('＜', '<')
    if o == r:
        return True
    if (o == '0' and r in ('未检出', '0')) or (r == '0' and o in ('未检出', '0')):
        return True
    return False


def check_original_records(registry, test_data):
    """Phase 1: Check original records for internal anomalies."""
    issues = []

    # Group samples by plant
    plant_samples = defaultdict(dict)
    for entry in registry:
        sid = entry['sample_id']
        if not sid.startswith('W'):
            continue
        desc = entry['description']
        wtype = classify_sample_water_type(desc)
        plant = extract_plant_from_desc(desc)
        if plant and wtype != '未知':
            plant_samples[plant][wtype] = sid

    # 1. Missing test data
    for entry in registry:
        sid = entry['sample_id']
        if sid not in test_data or not test_data[sid]:
            issues.append(f"样品 {sid}（{entry['description']}）在原始记录中无任何检测数据")

    # 2. pH range
    for sid, items in test_data.items():
        if not sid.startswith('W'):
            continue
        entry = next((e for e in registry if e['sample_id'] == sid), None)
        label = f"{entry['description']}({sid})" if entry else sid
        ph = items.get('pH')
        if ph:
            try:
                phv = float(str(ph).replace('<', ''))
                if phv < 5 or phv > 10:
                    issues.append(f"[严重] {label} pH={phv} 异常（通常范围 5-10）")
            except (ValueError, TypeError):
                pass

    # 3. Negative values
    for sid, items in test_data.items():
        entry = next((e for e in registry if e['sample_id'] == sid), None)
        label = f"{entry['description']}({sid})" if entry else sid
        for name, val in items.items():
            try:
                if float(str(val).replace('<', '').replace('＜', '')) < 0:
                    issues.append(f"[严重] {label} 项目「{name}」值为负数({val})")
            except (ValueError, TypeError):
                pass

    # 4. Chlorine: 管网水 should <= 出厂水
    for plant, sids in plant_samples.items():
        cc_sid = sids.get('出厂水')
        gw_sid = sids.get('管网水') or sids.get('管网末梢水')
        if not (cc_sid and gw_sid):
            continue
        cc_data, gw_data = test_data.get(cc_sid, {}), test_data.get(gw_sid, {})
        for cl in ['游离氯', '二氧化氯']:
            try:
                ccv = float(str(cc_data.get(cl, '')).replace('<', ''))
                gwv = float(str(gw_data.get(cl, '')).replace('<', ''))
                if gwv > ccv * 1.1 and gwv > 0 and ccv > 0:
                    issues.append(f"{plant} 管网水{cl}({gwv})高于出厂水({ccv})，需确认")
            except (ValueError, TypeError):
                pass

    # 5. KMnO4: 出厂水 should <= 原水
    for plant, sids in plant_samples.items():
        cc_sid, yw_sid = sids.get('出厂水'), sids.get('原水')
        if not (cc_sid and yw_sid):
            continue
        cc_data, yw_data = test_data.get(cc_sid, {}), test_data.get(yw_sid, {})
        for key in ['高锰酸盐指数', '高锰酸盐指数(以O2计)']:
            cc_v, yw_v = cc_data.get(key, ''), yw_data.get(key, '')
            if cc_v and yw_v:
                try:
                    ccf = float(str(cc_v).replace('<', ''))
                    ywf = float(str(yw_v).replace('<', ''))
                    if ccf > ywf * 1.2:
                        issues.append(f"{plant} 出厂水高锰酸盐指数({ccf})高于原水({ywf})，异常")
                except (ValueError, TypeError):
                    pass
                break

    # 6. Quality control samples (M/K series) negative check
    for sid, items in test_data.items():
        if not (sid.startswith('M') or sid.startswith('K')):
            continue
        for name, val in items.items():
            try:
                if float(str(val).replace('<', '').replace('＜', '')) < 0:
                    issues.append(f"[严重] 质控样品 {sid} 项目「{name}」值为负数({val})")
            except (ValueError, TypeError):
                pass

    # 7. Duplicate values detection (potential copy-paste errors)
    item_val_samples = defaultdict(lambda: defaultdict(list))
    for sid, items in test_data.items():
        if not sid.startswith('W'):
            continue
        for name, val in items.items():
            if val.startswith('<') or val.startswith('＜') or val in ('未检出', '无', '0'):
                continue
            try:
                float(val)
                if '.' in val and len(val.split('.')[1]) >= 3:
                    item_val_samples[name][val].append(sid)
            except ValueError:
                pass
    for name, val_map in item_val_samples.items():
        for val, sids in val_map.items():
            if len(sids) >= 4:
                descs = [next((e['description'] for e in registry if e['sample_id'] == s), s) for s in sids[:5]]
                issues.append(
                    f"项目「{name}」有 {len(sids)} 个样品结果完全相同({val})，"
                    f"涉及：{'、'.join(descs)}{'...' if len(sids) > 5 else ''}，请确认是否录入错误")

    # 8. Turbidity logic: 出厂水 should be lower than 原水
    for plant, sids in plant_samples.items():
        cc_sid, yw_sid = sids.get('出厂水'), sids.get('原水')
        if not (cc_sid and yw_sid):
            continue
        cc_turb = test_data.get(cc_sid, {}).get('浑浊度', '')
        yw_turb = test_data.get(yw_sid, {}).get('浑浊度', '')
        if cc_turb and yw_turb:
            try:
                ccf = float(str(cc_turb).replace('<', ''))
                ywf = float(str(yw_turb).replace('<', ''))
                if ccf > ywf and ywf > 0:
                    issues.append(f"{plant} 出厂水浑浊度({ccf})高于原水({ywf})，异常")
            except (ValueError, TypeError):
                pass

    # 9. Bacterial indicators: 出厂水/管网水 should be 0/未检出
    for sid, items in test_data.items():
        if not sid.startswith('W'):
            continue
        entry = next((e for e in registry if e['sample_id'] == sid), None)
        if not entry:
            continue
        wtype = classify_sample_water_type(entry['description'])
        if wtype not in ('出厂水', '管网水', '管网末梢水'):
            continue
        label = f"{entry['description']}({sid})"
        for bact in ['菌落总数', '总大肠菌群', '大肠埃希氏菌']:
            val = items.get(bact, '')
            if not val or val in ('0', '未检出', '<1'):
                continue
            try:
                if float(val) > 0:
                    issues.append(f"{label} {bact}={val}，出厂水/管网水该指标通常应为0或未检出")
            except (ValueError, TypeError):
                pass

    # 10. Same-source raw water consistency
    source_groups = defaultdict(list)
    for entry in registry:
        sid = entry['sample_id']
        if not sid.startswith('W') or '原水' not in entry['description']:
            continue
        m = re.search(r'[\(（]([^)）]+)[\)）]', entry['description'])
        source = m.group(1) if m else entry['description']
        source_groups[source].append((sid, entry))
    for source, entries in source_groups.items():
        if len(entries) < 2:
            continue
        for param in ['pH', '高锰酸盐指数', '溶解氧', '浑浊度']:
            vals = {}
            for sid, entry in entries:
                v = test_data.get(sid, {}).get(param)
                if v:
                    try:
                        vals[sid] = (float(str(v).replace('<', '')), entry['description'])
                    except (ValueError, TypeError):
                        pass
            if len(vals) >= 2:
                vlist = [v[0] for v in vals.values()]
                if min(vlist) > 0 and max(vlist) / min(vlist) > 2:
                    detail = ', '.join(f"{desc}={v}" for _, (v, desc) in vals.items())
                    issues.append(f"同源原水「{source}」{param}差异较大：{detail}")

    # 11. Logical consistency checks per sample
    for sid, items in test_data.items():
        if not sid.startswith('W'):
            continue
        entry = next((e for e in registry if e['sample_id'] == sid), None)
        label = f"{entry['description']}({sid})" if entry else sid
        issues.extend(check_data_logic(items, label))

    # 12. Significant figures consistency within original records (grouped by water type)
    wtype_items = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for entry in registry:
        sid = entry['sample_id']
        if not sid.startswith('W') or sid not in test_data:
            continue
        wtype = classify_sample_water_type(entry['description'])
        for param, val in test_data[sid].items():
            s = str(val).strip().replace('＜', '<')
            if not s or s.startswith('<') or s in ('未检出', '无', '0'):
                continue
            try:
                fval = float(s)
                if fval == 0:
                    continue
            except (ValueError, TypeError):
                continue
            sf = count_digits(s)
            wtype_items[wtype][param][sf].append((sid, entry['description'], s))
    for wtype, params in wtype_items.items():
        for param, sf_map in params.items():
            if len(sf_map) <= 1:
                continue
            most = max(sf_map.items(), key=lambda x: len(x[1]))
            for sf, entries in sf_map.items():
                if sf != most[0] and len(entries) < len(most[1]):
                    samples = ', '.join(f"{desc}(样品编号{sid})={val}" for sid, desc, val in entries[:3])
                    suffix = '...' if len(entries) > 3 else ''
                    issues.append(
                        f"原始记录({wtype})「{param}」数字位数不一致："
                        f"{len(entries)}个样品为{sf}位，"
                        f"多数({len(most[1])})为{most[0]}位，"
                        f"涉及：{samples}{suffix}")

    return issues


def cross_verify_reports(registry, test_data, all_info):
    """Cross-verify reports against original records."""
    issues_verify = []
    issues_logic = []

    # Build sample_id -> report mapping
    sid_to_report = {}
    for fname, info in all_info.items():
        sid = info.get('sample_id', '').strip()
        if sid:
            sid_to_report[sid] = (fname, info)

    # A. Registry samples missing reports
    for entry in registry:
        sid = entry['sample_id']
        if not sid.startswith('W'):
            continue
        if sid not in sid_to_report:
            issues_verify.append(
                f"原始记录样品 {sid}（{entry['description']}）未找到对应报告文件")

    # B. Report sample_id not in registry
    known_sids = {e['sample_id'] for e in registry}
    for fname, info in all_info.items():
        sid = info.get('sample_id', '').strip()
        if sid and sid not in known_sids and re.match(r'[WKM]\d{6}C\d+', sid):
            issues_verify.append(
                f"报告 \"{fname}\" 样品编号「{sid}」在原始记录中无对应，可能属于其他批次")

    # C. Company name consistency
    for entry in registry:
        sid = entry['sample_id']
        if sid not in sid_to_report:
            continue
        fname, info = sid_to_report[sid]
        expected, actual = entry['company'], info.get('company', '')
        if expected and actual and expected not in actual and actual not in expected:
            issues_verify.append(
                f"报告 \"{fname}\" 被检单位不一致：原始记录「{expected}」，报告「{actual}」")

    # D. Test data value comparison
    for entry in registry:
        sid = entry['sample_id']
        if sid not in sid_to_report or sid not in test_data:
            continue
        fname, info = sid_to_report[sid]
        orig_items = test_data[sid]
        report_items = info.get('test_items', [])
        is_raw = '原水' in entry['description']

        for orig_name, orig_val in orig_items.items():
            # Extract base name (before any paren/space+unit) for skip check
            base_orig = orig_name.split('(')[0].split('（')[0].strip()
            base_orig = base_orig.split(' ')[0] if ' ' in base_orig else base_orig
            if base_orig in ('钙', '镁', '电导率', '水温'):
                continue
            if is_raw and orig_name in ('肉眼可见物', '臭和味'):
                continue
            matched = find_matching_report_item(report_items, orig_name)
            if matched is None:
                issues_verify.append(
                    f"报告 \"{fname}\" 缺少原始记录项目「{orig_name}」(原始值={orig_val})")
                continue
            if not vals_match(orig_val, matched['result']):
                o = re.sub(r'[、，,]+$', '', str(orig_val).strip()).replace('＜', '<')
                r = re.sub(r'[、，,]+$', '', str(matched['result']).strip()).replace('＜', '<')
                # Distinguish value difference vs formatting/sig-fig difference
                try:
                    ov, rv = float(o.replace('<', '')), float(r.replace('<', ''))
                    is_value_diff = abs(ov - rv) > 0.0001
                except (ValueError, TypeError):
                    is_value_diff = True
                if is_value_diff:
                    tag = "[严重-数值] "
                    detail = "数值不一致"
                else:
                    tag = "[严重-位数] "
                    detail = "数字位数不一致"
                issues_verify.append(
                    f"{tag}报告 \"{fname}\" {detail} - 「{orig_name}」: "
                    f"原始记录={orig_val}, 报告={matched['result']}")

    # E. Cross-report logic: group by plant
    plant_reports = defaultdict(dict)
    for entry in registry:
        sid = entry['sample_id']
        if sid not in sid_to_report:
            continue
        wtype = classify_sample_water_type(entry['description'])
        plant = extract_plant_from_desc(entry['description'])
        if plant and wtype != '未知':
            plant_reports[plant][wtype] = sid_to_report[sid]

    for plant, tm in plant_reports.items():
        # Chlorine: 管网 <= 出厂
        cc = tm.get('出厂水')
        gw = tm.get('管网水') or tm.get('管网末梢水')
        if cc and gw:
            for cl in ['游离氯', '二氧化氯']:
                cc_i = find_matching_report_item(cc[1].get('test_items', []), cl)
                gw_i = find_matching_report_item(gw[1].get('test_items', []), cl)
                if cc_i and gw_i:
                    try:
                        ccv = float(cc_i['result'].replace('<', '').replace('＜', ''))
                        gwv = float(gw_i['result'].replace('<', '').replace('＜', ''))
                        if gwv > ccv * 1.1 and ccv > 0:
                            issues_logic.append(
                                f"{plant} 管网水({gw[0]}){cl}({gwv})高于出厂水({cc[0]})({ccv})")
                    except (ValueError, TypeError):
                        pass

        # KMnO4: 出厂 <= 原水
        yw = tm.get('原水')
        if cc and yw:
            for kn in ['高锰酸盐指数(以O2计)', '高锰酸盐指数']:
                cc_i = find_matching_report_item(cc[1].get('test_items', []), kn)
                yw_i = find_matching_report_item(yw[1].get('test_items', []), kn)
                if cc_i and yw_i:
                    try:
                        ccv = float(cc_i['result'].replace('<', '').replace('＜', ''))
                        ywv = float(yw_i['result'].replace('<', '').replace('＜', ''))
                        if ccv > ywv * 1.5:
                            issues_logic.append(
                                f"{plant} 出厂水({cc[0]})高锰酸盐指数({ccv})"
                                f"显著高于原水({yw[0]})({ywv})")
                    except (ValueError, TypeError):
                        pass
                    break

    # F. Logical consistency per report
    for fname, info in all_info.items():
        items_dict = {item['name']: item['result'] for item in info.get('test_items', [])}
        if items_dict:
            issues_logic.extend(check_data_logic(items_dict, f"报告\"{fname}\""))

    # G. Conclusion vs actual data - flag if "合格/符合" but has exceedances
    for fname, info in all_info.items():
        conclusion = info.get('conclusion', '')
        if not conclusion or '不' in conclusion:
            continue
        if not ('符合' in conclusion or '合格' in conclusion or '满足' in conclusion):
            continue
        for item in info.get('test_items', []):
            result = item['result'].replace('＜', '<')
            standard = item['standard']
            if not result or result.startswith('<') or '水温' in item['name']:
                continue
            if result in ('无', '未检出', '无异臭、异味', '0'):
                continue
            try:
                val = float(result)
            except (ValueError, TypeError):
                continue
            std_m = re.search(r'[≤<]\s*([\d.]+)', standard)
            if std_m:
                try:
                    if val > float(std_m.group(1)):
                        issues_verify.append(
                            f"报告 \"{fname}\" 结论为「{conclusion[:50]}」，"
                            f"但「{item['name']}」结果({result})超标({standard})，结论与数据矛盾")
                        break
                except (ValueError, TypeError):
                    pass

    # G. Raw water standard reference
    for fname, info in all_info.items():
        if info.get('water_type') != '原水':
            continue
        std = info.get('product_standard', '')
        conclusion = info.get('conclusion', '')
        if '生活饮用水' in std or '5749' in std:
            issues_verify.append(
                f"报告 \"{fname}\" 为原水报告但引用了生活饮用水标准，通常应引用地表水标准（GB 3838）")
        elif conclusion and '生活饮用水' in conclusion:
            issues_verify.append(
                f"报告 \"{fname}\" 为原水报告但结论中引用了生活饮用水标准，请确认")

    # H. Sampling location consistency
    for entry in registry:
        sid = entry['sample_id']
        if sid not in sid_to_report:
            continue
        fname, info = sid_to_report[sid]
        reg_desc = entry['description']
        report_loc = info.get('sampling_location', '')
        if not (reg_desc and report_loc):
            continue
        m = re.search(r'[\(（]([^)）]+)[\)）]', reg_desc)
        if m:
            key_loc = m.group(1)
            plant = extract_plant_from_desc(reg_desc)
            if key_loc not in report_loc and report_loc not in key_loc:
                if plant and plant not in report_loc:
                    issues_verify.append(
                        f"报告 \"{fname}\" 采样地点可能不一致：原始记录「{reg_desc}」，"
                        f"报告「{report_loc}」")

    # I. Testing method consistency across same-type reports
    type_methods = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for fname, info in all_info.items():
        wt = info.get('water_type', '')
        for item in info.get('test_items', []):
            if item.get('method'):
                type_methods[wt][item['name']][normalize_method(item['method'])].append(fname)
    for wt, items_map in type_methods.items():
        for item_name, methods in items_map.items():
            if len(methods) <= 1:
                continue
            most_common = max(methods.items(), key=lambda x: len(x[1]))
            for method, fnames in methods.items():
                if method != most_common[0] and len(fnames) < len(most_common[1]):
                    issues_logic.append(
                        f"同类型({wt})报告「{item_name}」检测方法不一致："
                        f"{len(fnames)}个使用「{method}」，"
                        f"多数({len(most_common[1])})使用「{most_common[0]}」，"
                        f"涉及：{', '.join(fnames[:3])}{'...' if len(fnames) > 3 else ''}")

    return issues_verify, issues_logic


# ─────────────────── MAIN ANALYSIS ───────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="水质检测报告验证分析脚本 —— 扫描指定目录下的 .xlsx/.xls 报告文件并输出待确认问题清单。",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""示例:
  %(prog)s                          # 扫描脚本所在目录
  %(prog)s /path/to/reports         # 扫描指定目录
  %(prog)s ./report/0189-0211       # 扫描相对路径目录
  %(prog)s -r /path/to/reports      # 同上（使用 -r 参数）
  %(prog)s -o result.txt            # 自定义输出文件名
""")
    parser.add_argument('directory', nargs='?', default=None,
                        help='报告文件所在目录（默认为脚本所在目录）')
    parser.add_argument('-r', '--report-dir', default=None,
                        help='报告文件所在目录（与位置参数等效，位置参数优先）')
    parser.add_argument('-o', '--output', default=None,
                        help='输出文件路径（默认为扫描目录下的"待确认问题清单.txt"）')
    parser.add_argument('-ori', action='store_true',
                        help='仅检查原始记录，不检查报告文件')
    return parser.parse_args()


def main():
    args = parse_args()

    # 确定扫描目录：位置参数 > -r 参数 > 脚本所在目录
    report_dir = args.directory or args.report_dir or os.path.dirname(os.path.abspath(__file__))
    report_dir = os.path.abspath(report_dir)

    if not os.path.isdir(report_dir):
        print(f"错误：目录不存在 —— {report_dir}")
        sys.exit(1)

    output_file = args.output or os.path.join(report_dir, "待确认问题清单.txt")
    output_file = os.path.abspath(output_file)

    # ══════════════════════════════════════════════════════
    # Phase 1: Find and read original record
    # ══════════════════════════════════════════════════════
    orig_record_file = find_original_record_file(report_dir)

    # 原始记录文件名，需从报告列表中排除
    orig_basename = os.path.basename(orig_record_file) if orig_record_file else None
    files = sorted([f for f in os.listdir(report_dir)
                    if f.endswith(('.xlsx', '.xls')) and not f.startswith('~')
                    and f != orig_basename])

    print(f"扫描目录：{report_dir}")
    print(f"共找到 {len(files)} 个报告文件，开始分析...")

    if not files:
        print("\n未��到任何 .xlsx / .xls 报告文件，请将报告文件放入扫描目录后重试。")
        print(f"扫描目录：{report_dir}")
        return
    registry, orig_test_data = [], {}
    issues_original = []

    if orig_record_file:
        orig_fname = os.path.basename(orig_record_file)
        print(f"找到原始记录文件：{orig_fname}")
        try:
            registry, orig_test_data = read_original_record(orig_record_file)
            print(f"  样品登记：{len(registry)} 条")
            print(f"  检测数据：{len(orig_test_data)} 个样品")
            print("正在检查原始记录...")
            issues_original = check_original_records(registry, orig_test_data)
            print(f"  原始记录问题：{len(issues_original)} 项")
        except Exception as e:
            print(f"  原始记录读取失败：{e}")
            traceback.print_exc()
    else:
        print("未找到原始记录文件（如 260205-1-25.xlsx），跳过原始记录检查与交叉验证。")
    print()

    # ── -ori 模式：仅输出原始记录检查结果 ──
    if args.ori:
        ori_output = output_file if args.output else os.path.join(report_dir, "原始记录检查结果.txt")
        lines = []
        lines.append("=" * 72)
        lines.append("    原始记录检查 —— 问题清单")
        lines.append("=" * 72)
        lines.append(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"扫描目录：{report_dir}")
        if orig_record_file:
            lines.append(f"原始记录：{os.path.basename(orig_record_file)}")
        lines.append("")
        lines.append(f"共发现 {len(issues_original)} 项待确认问题")
        lines.append("-" * 72)
        if not issues_original:
            lines.append("  （无）")
        for i, issue in enumerate(issues_original, 1):
            lines.append(f"  {i}. {issue}")
        lines.append("")
        lines.append("=" * 72)
        lines.append("以上问题均为程序自动检测，可能存在误报，请人工逐项核实。")
        lines.append("=" * 72)

        output_text = '\n'.join(lines)
        with open(ori_output, 'w', encoding='utf-8') as f:
            f.write(output_text)

        print(f"\n原始记录检查完成！共发现 {len(issues_original)} 项待确认问题。")
        print(f"结果已写入：{ori_output}")
        return output_text

    # ══════════════════════════════════════════════════════
    # Phase 2: Read and check report files
    # ══════════════════════════════════════════════════════

    # ── Collect all file info ──
    all_info = {}  # fname -> info dict
    for i, fname in enumerate(files):
        filepath = os.path.join(report_dir, fname)
        if fname.endswith('.xlsx'):
            info = read_xlsx_report_info(filepath)
        else:
            info = read_xls_report_info(filepath)
        info['filename'] = fname
        info['extension'] = os.path.splitext(fname)[1]
        info['prefix'] = extract_number_prefix(fname)
        info['water_type'] = classify_water_type(fname)
        info['plant_name'] = extract_plant_name(fname)
        all_info[fname] = info
        if (i + 1) % 20 == 0:
            print(f"  已处理 {i+1}/{len(files)} ...")

    print(f"文件读取完成，开始问题检测...")

    # ════════════════════════════════════════════════════
    # Issue categories
    # ════════════════════════════════════════════════════
    issues_naming = []       # 一、命名问题
    issues_numbering = []    # 二、编号问题
    issues_data = []         # 三、数据问题
    issues_format = []       # 四、格式/模板问题
    issues_date = []         # 五、日期问题
    issues_consistency = []  # 六、一致性问题
    issues_values = []       # 七、异常值问题
    issues_read_errors = []  # 八、文件读取问题

    # ──────────── 一、命名问题 ────────────
    # 1. Check bracket matching in filenames
    for fname in files:
        name_part = os.path.splitext(fname)[0]
        open_cn = name_part.count('（')
        close_cn = name_part.count('）')
        open_en = name_part.count('(')
        close_en = name_part.count(')')
        if open_cn != close_cn:
            issues_naming.append(f"文件 \"{fname}\" 中文括号不匹配：'（' 出现 {open_cn} 次，'）' 出现 {close_cn} 次")
        if open_en != close_en:
            issues_naming.append(f"文件 \"{fname}\" 英文括号不匹配：'(' 出现 {open_en} 次，')' 出现 {close_en} 次")

    # 2. Check for inconsistent numbering prefix length
    prefixes = [(fname, extract_number_prefix(fname)) for fname in files]
    prefix_lengths = Counter(len(p) for _, p in prefixes if p)
    if len(prefix_lengths) > 1:
        for fname, p in prefixes:
            if p and len(p) != 4:
                issues_naming.append(f"文件 \"{fname}\" 编号前缀位数异常：'{p}' 为 {len(p)} 位，多数文件为 4 位")

    # 3. Number sequence gaps (only within actual file range)
    nums = sorted(set(int(p) for _, p in prefixes if p))
    expected = set(range(min(nums), max(nums) + 1))
    actual = set(nums)
    missing = sorted(expected - actual)
    if missing:
        # Group consecutive missing numbers
        groups = []
        start = missing[0]
        end = missing[0]
        for n in missing[1:]:
            if n == end + 1:
                end = n
            else:
                groups.append((start, end))
                start = n
                end = n
        groups.append((start, end))
        for s, e in groups:
            if s == e:
                issues_naming.append(f"文件编号序列缺失：{s:04d}")
            else:
                issues_naming.append(f"文件编号序列缺失：{s:04d}-{e:04d}（共 {e-s+1} 个）")

    # 4. Duplicate prefix numbers
    prefix_counter = Counter(int(p) for _, p in prefixes if p)
    for num, cnt in sorted(prefix_counter.items()):
        if cnt > 1:
            dup_files = [f for f, p in prefixes if p and int(p) == num]
            issues_naming.append(f"文件编号重复：编号 {num:04d} 出现 {cnt} 次，涉及文件：{', '.join(dup_files)}")

    # 5. Inconsistent naming patterns
    # Check for extra spaces in filenames
    for fname in files:
        if '  ' in fname:
            issues_naming.append(f"文件 \"{fname}\" 名称中包含连续空格")
        if fname != fname.strip():
            issues_naming.append(f"文件 \"{fname}\" 名称首尾有多余空格")

    # 6. Check for "水厂水厂" duplicated in name
    for fname in files:
        if '水厂水厂' in fname:
            issues_naming.append(f"文件 \"{fname}\" 名称中 '水厂' 重复出现，可能为笔误")

    # 7. Check for inconsistent date suffixes in filenames
    # Some files have date suffix like "01.05" and some don't
    files_with_date = [f for f in files if re.search(r'\d{2}\.\d{2}\.(xlsx?|xls)$', f) or
                       re.search(r'\d{2}\.\d{2}\s*\.(xlsx?|xls)$', f)]
    files_without_date = [f for f in files if f not in files_with_date]
    if files_with_date and files_without_date and len(files_with_date) < len(files_without_date):
        issues_naming.append(
            f"部分文件名含日期后缀（共 {len(files_with_date)} 个），"
            f"而多数文件不含日期后缀（共 {len(files_without_date)} 个），格式不统一。"
            f"含日期的文件：{', '.join(files_with_date)}")

    # 8. Check extension vs water type consistency
    # Expectation: 原水 -> .xls, 出厂水/管网水 -> .xlsx (general pattern)
    for fname, info in all_info.items():
        wt = info['water_type']
        ext = info['extension']
        if wt == '原水' and ext == '.xlsx':
            issues_naming.append(f"文件 \"{fname}\" 为原水报告但使用 .xlsx 格式，"
                                 f"一般原水报告使用 .xls 格式，请确认")
        if wt in ('出厂水',) and ext == '.xls':
            issues_naming.append(f"文件 \"{fname}\" 为出厂水报告但使用 .xls 格式，"
                                 f"一般出厂水报告使用 .xlsx 格式，请确认")
        # 管网水 can be either, but check for xls ones (pattern: 管网水 xls typically only for early files with brackets)
        if wt == '二次供水' and ext == '.xlsx':
            issues_naming.append(f"文件 \"{fname}\" 为二次供水报告但使用 .xlsx 格式，"
                                 f"一般二次供水报告使用 .xls 格式，请确认")

    # 9. Check for inconsistent water type labeling in filename
    # Two patterns exist: "水厂（管网水）" and "水厂管网水" -- flag the inconsistency as a whole
    guanwang_files = [f for f in files if '管网' in f]
    guanwang_bracket = [f for f in guanwang_files if '（管网' in f or '（管网' in f]
    guanwang_no_bracket = [f for f in guanwang_files if f not in guanwang_bracket]
    if guanwang_bracket and guanwang_no_bracket:
        issues_naming.append(
            f"管网水文件命名格式不统一：{len(guanwang_bracket)} 个文件使用括号形式如 '水厂（管网水）'，"
            f"{len(guanwang_no_bracket)} 个文件使用无括号形式如 '水厂管网水'，建议统一命名规范")

    # ──────────── 二、编号问题 ────────────
    # Check report numbers from file content vs filename prefix
    report_nums = {}  # report_number -> [filenames]
    for fname, info in all_info.items():
        rn = info.get('report_number')
        if rn:
            report_nums.setdefault(rn, []).append(fname)
            # Compare with filename prefix
            prefix = info.get('prefix')
            if prefix:
                # Normalize: remove leading zeros for comparison
                try:
                    if int(rn) != int(prefix):
                        issues_numbering.append(
                            f"文件 \"{fname}\" 的文件名编号 ({prefix}) 与报告内编号 ({rn}) 不一致")
                except ValueError:
                    issues_numbering.append(
                        f"文件 \"{fname}\" 的报告内编号 '{rn}' 无法解析为数字")

    # Check for duplicate report numbers in content
    for rn, fnames in sorted(report_nums.items(), key=lambda x: x[0]):
        if len(fnames) > 1:
            issues_numbering.append(
                f"报告编号 {rn} 在多个文件中重复使用：{', '.join(fnames)}")

    # Check for missing report numbers
    for fname, info in all_info.items():
        if 'report_number' not in info and 'read_error' not in info:
            issues_numbering.append(f"文件 \"{fname}\" 未能提取到报告编号")

    # ──────────── 三、数据问题 ────────────
    for fname, info in all_info.items():
        # Missing sample name
        if 'sample_name' not in info and 'read_error' not in info:
            issues_data.append(f"文件 \"{fname}\" 未提取到样品名称")

        # Missing company
        if 'company' not in info and 'read_error' not in info:
            issues_data.append(f"文件 \"{fname}\" 未提取到被检单位名称")

        # Missing sampling date
        if 'sampling_date' not in info and 'read_error' not in info:
            issues_data.append(f"文件 \"{fname}\" 未提取到采样日期")

        # Missing sample ID
        if 'sample_id' not in info and 'read_error' not in info:
            issues_data.append(f"文件 \"{fname}\" 未提取到样品编号")

        # Check test items for blank results
        test_items = info.get('test_items', [])
        blank_items = [item['name'] for item in test_items if not item['result'] or item['result'] == 'None']
        if blank_items:
            issues_data.append(
                f"文件 \"{fname}\" 以下检测项目结果为空：{', '.join(blank_items)}")

        # Check test items for blank methods
        no_method_items = [item['name'] for item in test_items if not item['method'] or item['method'] == 'None']
        if no_method_items:
            issues_data.append(
                f"文件 \"{fname}\" 以下检测项目缺少检测方法：{', '.join(no_method_items)}")

        # Check declared test item count vs actual
        declared = info.get('test_item_count')
        actual_count = len(test_items)
        if declared and actual_count > 0:
            if actual_count != declared:
                issues_data.append(
                    f"文件 \"{fname}\" 声称检测 {declared} 项指标，"
                    f"但实际提取到 {actual_count} 项数据")

        # Missing conclusion
        if 'conclusion' not in info and 'read_error' not in info:
            issues_data.append(f"文件 \"{fname}\" 未提取到检测结论")

        # Missing report date
        if 'report_date' not in info and 'read_error' not in info:
            issues_data.append(f"文件 \"{fname}\" 未提取到报告编制日期")

        # Duplicate test items within a single report
        item_names = [item['name'] for item in test_items]
        name_counts = Counter(item_names)
        for name, cnt in name_counts.items():
            if cnt > 1:
                issues_data.append(f"文件 \"{fname}\" 检测项目「{name}」重复出现 {cnt} 次")

    # ──────────── 数字位数一致性 ────────────
    # Group by water type (also used later for format checks)
    type_groups = defaultdict(list)
    for fname, info in all_info.items():
        wt = info['water_type']
        type_groups[wt].append((fname, info))

    for wt, group in type_groups.items():
        if len(group) < 2:
            continue
        item_sigfigs = defaultdict(lambda: defaultdict(list))
        for fname, info in group:
            for item in info.get('test_items', []):
                result = item['result']
                if not result:
                    continue
                if result.startswith('<') or result.startswith('＜'):
                    continue
                if result in ('未检出', '无', '无异臭、异味', '0'):
                    continue
                try:
                    fval = float(result)
                    if fval == 0:
                        continue
                except (ValueError, TypeError):
                    continue
                sf = count_digits(result)
                item_sigfigs[item['name']][sf].append((fname, result))
        for item_name, sf_map in item_sigfigs.items():
            if len(sf_map) <= 1:
                continue
            most = max(sf_map.items(), key=lambda x: len(x[1]))
            for sf, entries in sf_map.items():
                if sf != most[0] and len(entries) < len(most[1]):
                    details = ', '.join(f"{fn}(值={val})" for fn, val in entries[:3])
                    suffix = '...' if len(entries) > 3 else ''
                    issues_data.append(
                        f"同类型({wt})报告「{item_name}」数字位数不一致："
                        f"{len(entries)}个报告为{sf}位数字，"
                        f"多数({len(most[1])})为{most[0]}位数字，"
                        f"涉及：{details}{suffix}")

    # ──────────── 四、格式/模板问题 ────────────
    for wt, group in type_groups.items():
        if len(group) < 2:
            continue

        # Compare sheet counts
        sheet_counts = Counter(info.get('sheet_count', 0) for _, info in group)
        if len(sheet_counts) > 1:
            most_common_count = sheet_counts.most_common(1)[0][0]
            for fname, info in group:
                sc = info.get('sheet_count', 0)
                if sc != most_common_count:
                    issues_format.append(
                        f"文件 \"{fname}\"（{wt}类）共 {sc} 个工作表，"
                        f"而同类报告多数为 {most_common_count} 个工作表")

        # Compare total pages
        page_counts = Counter(info.get('total_pages', 0) for _, info in group)
        if len(page_counts) > 1:
            most_common_pages = page_counts.most_common(1)[0][0]
            for fname, info in group:
                pc = info.get('total_pages', 0)
                if pc != most_common_pages and pc != 0:
                    issues_format.append(
                        f"文件 \"{fname}\"（{wt}类）报告页数为 {pc} 页，"
                        f"而同类报告多数为 {most_common_pages} 页")

        # Compare product standards
        standards = Counter(info.get('product_standard', '未知') for _, info in group)
        if len(standards) > 1:
            most_common_std = standards.most_common(1)[0][0]
            for fname, info in group:
                std = info.get('product_standard', '未知')
                if std != most_common_std and std != '未知':
                    issues_format.append(
                        f"文件 \"{fname}\"（{wt}类）产品标准为 \"{std}\"，"
                        f"而同类报告多数为 \"{most_common_std}\"")

        # Compare test item counts
        item_counts = [info.get('test_item_count', 0) for _, info in group if info.get('test_item_count')]
        if item_counts:
            common_count = Counter(item_counts).most_common(1)[0][0]
            for fname, info in group:
                ic = info.get('test_item_count', 0)
                if ic and ic != common_count:
                    issues_format.append(
                        f"文件 \"{fname}\"（{wt}类）检测项目数为 {ic} 项，"
                        f"而同类报告多数为 {common_count} 项")

    # Check if 管网水 xls files follow a different template from xlsx ones
    guanwang_xls = [(f, i) for f, i in all_info.items()
                    if i['water_type'] == '管网水' and i['extension'] == '.xls']
    guanwang_xlsx = [(f, i) for f, i in all_info.items()
                     if i['water_type'] == '管网水' and i['extension'] == '.xlsx']
    if guanwang_xls and guanwang_xlsx:
        xls_pages = Counter(i.get('total_pages', 0) for _, i in guanwang_xls)
        xlsx_pages = Counter(i.get('total_pages', 0) for _, i in guanwang_xlsx)
        issues_format.append(
            f"管网水报告中，.xls 文件共 {len(guanwang_xls)} 个（页数分布：{dict(xls_pages)}），"
            f".xlsx 文件共 {len(guanwang_xlsx)} 个（页数分布：{dict(xlsx_pages)}），请确认是否使用不同模板")

    # ──────────── 页码检查 ────────────
    for fname, info in all_info.items():
        filepath = os.path.join(report_dir, fname)
        total_sheets = info.get('sheet_count', 0)
        if total_sheets == 0:
            continue
        try:
            if fname.endswith('.xlsx'):
                wb = openpyxl.load_workbook(filepath, data_only=True)
                sheet_names = wb.sheetnames
                for si, sn in enumerate(sheet_names):
                    ws = wb[sn]
                    page_found = False
                    for r in range(1, min(3, ws.max_row + 1)):
                        for c in range(1, min(ws.max_column + 1, 10)):
                            v = ws.cell(r, c).value
                            if v is None:
                                continue
                            s = str(v)
                            m = re.search(r'第\s*(\d+)\s*页\s*共\s*(\d+)\s*页', s)
                            if m:
                                page_found = True
                                page_num = int(m.group(1))
                                page_total = int(m.group(2))
                                expected_page = si + 1
                                if page_num != expected_page:
                                    issues_format.append(
                                        f"文件 \"{fname}\" 第{expected_page}个工作表页码标注为"
                                        f"\"第 {page_num} 页\"，应为\"第 {expected_page} 页\"")
                                if page_total != total_sheets:
                                    issues_format.append(
                                        f"文件 \"{fname}\" 第{expected_page}个工作表标注"
                                        f"\"共 {page_total} 页\"，实际共 {total_sheets} 页")
                                break
                        if page_found:
                            break
                    if not page_found:
                        issues_format.append(
                            f"文件 \"{fname}\" 第{si+1}个工作表未找到页码标注")
                wb.close()
            else:
                wb = xlrd.open_workbook(filepath)
                for si in range(wb.nsheets):
                    ws = wb.sheet_by_index(si)
                    page_found = False
                    for r in range(min(2, ws.nrows)):
                        for c in range(min(ws.ncols, 10)):
                            v = ws.cell_value(r, c)
                            if v in ('', None):
                                continue
                            s = str(v)
                            m = re.search(r'第\s*(\d+)\s*页\s*共\s*(\d+)\s*页', s)
                            if m:
                                page_found = True
                                page_num = int(m.group(1))
                                page_total = int(m.group(2))
                                expected_page = si + 1
                                if page_num != expected_page:
                                    issues_format.append(
                                        f"文件 \"{fname}\" 第{expected_page}个工作表页码标注为"
                                        f"\"第 {page_num} 页\"，应为\"第 {expected_page} 页\"")
                                if page_total != total_sheets:
                                    issues_format.append(
                                        f"文件 \"{fname}\" 第{expected_page}个工作表标注"
                                        f"\"共 {page_total} 页\"，实际共 {total_sheets} 页")
                                break
                        if page_found:
                            break
                    if not page_found:
                        issues_format.append(
                            f"文件 \"{fname}\" 第{si+1}个工作表未找到页码标注")
        except Exception:
            pass

    # Check sampler consistency within type groups
    for wt, group in type_groups.items():
        samplers = defaultdict(list)
        for fname, info in group:
            s = info.get('sampler', '未知')
            samplers[s].append(fname)
        if len(samplers) > 1 and '未知' not in samplers:
            # This is informational, not necessarily an issue
            pass

    # ──────────── 五、日期问题 ────────────
    for fname, info in all_info.items():
        # Check sampling date format
        sd = info.get('sampling_date', '')
        if sd:
            # Expected format: 2026.01.05
            if not re.match(r'^20\d{2}\.\d{2}\.\d{2}$', sd):
                issues_date.append(f"文件 \"{fname}\" 采样日期格式异常：'{sd}'")

        # Check receipt date vs sampling date
        rd = info.get('receipt_date', '')
        if sd and rd:
            # Receipt should be >= sampling, and at most 1 day apart
            try:
                sd_parsed = datetime.strptime(sd, '%Y.%m.%d')
                rd_parsed = datetime.strptime(rd, '%Y.%m.%d')
                diff = (rd_parsed - sd_parsed).days
                if diff < 0:
                    issues_date.append(
                        f"文件 \"{fname}\" 收样日期 ({rd}) 早于采样日期 ({sd})")
                elif diff > 1:
                    issues_date.append(
                        f"文件 \"{fname}\" 收样日期 ({rd}) 与采样日期 ({sd}) 间隔 {diff} 天，"
                        f"原则上应为同一天，最多不超过1天")
            except ValueError:
                pass

        # Check testing date
        td = info.get('testing_date', '')
        if td and sd:
            # Testing date format: 2026.01.05~01.16
            m = re.match(r'(20\d{2}\.\d{2}\.\d{2})~(\d{2}\.\d{2})', td)
            if m:
                try:
                    td_start = datetime.strptime(m.group(1), '%Y.%m.%d')
                    sd_parsed = datetime.strptime(sd, '%Y.%m.%d')
                    if td_start < sd_parsed:
                        issues_date.append(
                            f"文件 \"{fname}\" 检测开始日期 ({m.group(1)}) 早于采样日期 ({sd})")
                except ValueError:
                    pass

        # Check report date format
        rpt_date = info.get('report_date', '')
        if rpt_date:
            # Various formats: "2026年 01月23日", "2026 年1 月 26日", "2026年 2月6日"
            m = re.search(r'(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日', rpt_date)
            if m:
                try:
                    rpt_parsed = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                    if sd:
                        sd_parsed = datetime.strptime(sd, '%Y.%m.%d')
                        if rpt_parsed < sd_parsed:
                            issues_date.append(
                                f"文件 \"{fname}\" 报告编制日期 ({rpt_date.strip()}) 早于采样日期 ({sd})")
                    # Check if report date is in a reasonable range (2026)
                    if rpt_parsed.year != 2026:
                        issues_date.append(
                            f"文件 \"{fname}\" 报告编制日期年份为 {rpt_parsed.year}，非 2026 年")
                except ValueError:
                    issues_date.append(f"文件 \"{fname}\" 报告编制日期无法解析：'{rpt_date}'")
            elif rpt_date.strip():
                issues_date.append(f"文件 \"{fname}\" 报告编制日期格式异常：'{rpt_date}'")

    # ──────────── 六、一致性问题 ────────────
    # Group files by plant name and check consistency
    plant_groups = defaultdict(list)
    for fname, info in all_info.items():
        plant = info.get('plant_name', '')
        if plant:
            plant_groups[plant].append((fname, info))

    for plant, group in plant_groups.items():
        if len(group) < 2:
            continue

        # Check if sample_name references are consistent
        companies = set()
        for fname, info in group:
            c = info.get('company', '')
            if c:
                companies.add(c)
        if len(companies) > 1:
            issues_consistency.append(
                f"水厂 \"{plant}\" 的相关报告中被检单位名称不一致：{', '.join(companies)}，"
                f"涉及文件：{', '.join(f for f, _ in group)}")

    # Check for similar plant names that might be the same plant (typos)
    plant_names = list(plant_groups.keys())
    for i in range(len(plant_names)):
        for j in range(i + 1, len(plant_names)):
            a, b = plant_names[i], plant_names[j]
            # Check if one is substring of the other or differ by just "水厂"
            if a in b or b in a:
                if a != b and abs(len(a) - len(b)) <= 2:
                    issues_consistency.append(
                        f"水厂名称疑似重复/不一致：\"{a}\" 与 \"{b}\"，请确认是否为同一水厂")

    # Check sample_name vs filename consistency
    for fname, info in all_info.items():
        sn = info.get('sample_name', '')
        if sn and '【' in sn and '】' in sn:
            m = re.search(r'【(.+?)】', sn)
            if m:
                sample_plant = m.group(1)
                # Remove trailing info like /地表水
                sample_plant = sample_plant.split('/')[0].strip()
                file_plant = info.get('plant_name', '')
                # Simple check: the sample plant name should appear in filename
                name_no_ext = os.path.splitext(fname)[0]
                prefix_removed = re.sub(r'^\d+', '', name_no_ext)
                if sample_plant not in prefix_removed and file_plant not in sample_plant:
                    # More lenient check
                    if sample_plant.replace('水厂', '') not in prefix_removed:
                        issues_consistency.append(
                            f"文件 \"{fname}\" 内样品名称为 \"{sn}\"，"
                            f"与文件名中的水厂名称不一致")

    # Check sample_type vs filename water type
    for fname, info in all_info.items():
        st = info.get('sample_type', '')
        wt = info['water_type']
        if st and wt != '未知':
            if wt == '出厂水' and '出厂水' not in st:
                issues_consistency.append(
                    f"文件 \"{fname}\" 文件名标注为出厂水，但内容样品类型为 \"{st}\"")
            elif wt == '原水' and '原水' not in st:
                issues_consistency.append(
                    f"文件 \"{fname}\" 文件名标注为原水，但内容样品类型为 \"{st}\"")
            elif wt == '管网水' and '管网' not in st:
                issues_consistency.append(
                    f"文件 \"{fname}\" 文件名标注为管网水，但内容样品类型为 \"{st}\"")
            elif wt == '二次供水' and '二次供水' not in st:
                issues_consistency.append(
                    f"文件 \"{fname}\" 文件名标注为二次供水，但内容样品类型为 \"{st}\"")

    # Check for plants that have 出厂水 but no 原水 or vice versa
    plant_types = defaultdict(set)
    for fname, info in all_info.items():
        plant = info.get('plant_name', '')
        wt = info['water_type']
        if plant and wt in ('出厂水', '原水', '管网水'):
            plant_types[plant].add(wt)
    for plant, types in sorted(plant_types.items()):
        if '出厂水' in types and '原水' not in types:
            issues_consistency.append(
                f"水厂 \"{plant}\" 有出厂水报告但缺少原水报告")
        if '原水' in types and '出厂水' not in types:
            # Could be just a water source, not necessarily an issue
            # Only flag if it looks like a water plant (not a reservoir)
            if '水厂' in plant:
                issues_consistency.append(
                    f"水厂 \"{plant}\" 有原水报告但缺少出厂水报告")

    # ──────────── 七、异常值问题 ────────────
    issues_values_critical = []  # Serious exceedances (>2x standard or toxic indicators)
    issues_values_normal = []    # Normal exceedances
    for fname, info in all_info.items():
        test_items = info.get('test_items', [])
        for item in test_items:
            result = item['result']
            name = item['name']
            standard = item['standard']

            if not result or result == 'None':
                continue

            # Skip water temperature -- the standard describes temperature CHANGE limits,
            # not absolute temperature. Absolute values like 14C are completely normal.
            if '水温' in name:
                continue

            # Try to parse numeric results
            numeric_result = None
            if result.startswith('<') or result.startswith('＜'):
                # Below detection limit - generally OK
                continue
            try:
                numeric_result = float(result)
            except (ValueError, TypeError):
                if result in ('无', '未检出', '无异臭、异味', '0', '0.0'):
                    continue
                continue

            if numeric_result is not None:
                exceeded = False
                std_limit_val = None

                if standard:
                    # Pattern: "≤X(II类)" or "≤X"
                    std_match = re.search(r'[≤<]\s*([\d.]+)', standard)
                    if std_match:
                        try:
                            std_limit_val = float(std_match.group(1))
                            if numeric_result > std_limit_val:
                                exceeded = True
                        except ValueError:
                            pass
                    # Pattern: just a number as limit
                    elif re.match(r'^[\d.]+$', str(standard)):
                        try:
                            std_limit_val = float(standard)
                            if numeric_result > std_limit_val:
                                exceeded = True
                        except ValueError:
                            pass
                    # Pattern: range like "0.1~0.8" or "0.02-0.8"
                    range_match = re.match(r'([\d.]+)\s*[~\-～]\s*([\d.]+)', standard)
                    if range_match:
                        try:
                            lo = float(range_match.group(1))
                            hi = float(range_match.group(2))
                            if numeric_result < lo or numeric_result > hi:
                                exceeded = True
                                std_limit_val = hi
                        except ValueError:
                            pass

                if exceeded:
                    # Determine severity
                    ratio_str = ""
                    is_critical = False
                    if std_limit_val and std_limit_val > 0:
                        ratio = numeric_result / std_limit_val
                        ratio_str = f"（为标准限值的 {ratio:.1f} 倍）"
                        if ratio >= 2.0:
                            is_critical = True
                    # Toxic heavy metals are always critical
                    if name in ('铅', '汞', '镉', '砷', '铬(六价)') and exceeded:
                        is_critical = True

                    msg = (f"文件 \"{fname}\" 检测项目 \"{name}\" "
                           f"结果 {result} 超出标准限值 {standard} {ratio_str}")
                    if is_critical:
                        issues_values_critical.append("[严重] " + msg)
                    else:
                        issues_values_normal.append(msg)

                # Check for suspicious values
                if name == 'pH' and numeric_result is not None:
                    if numeric_result < 5 or numeric_result > 10:
                        issues_values_critical.append(
                            f"[严重] 文件 \"{fname}\" pH 值 {result} 异常（通常范围 6-9）")

                # Negative values
                if numeric_result < 0:
                    issues_values_critical.append(
                        f"[严重] 文件 \"{fname}\" 检测项目 \"{name}\" 结果为负值 {result}")

    issues_values = issues_values_critical + issues_values_normal

    # ──────────── 八、文件读取问题 ────────────
    for fname, info in all_info.items():
        if 'read_error' in info:
            issues_read_errors.append(f"文件 \"{fname}\" 读取异常：{info['read_error']}")

    # ──────────── 交叉验证 ────────────
    issues_cross_verify = []
    issues_cross_logic = []
    if registry and orig_test_data:
        print("正在进行交叉验证...")
        issues_cross_verify, issues_cross_logic = cross_verify_reports(
            registry, orig_test_data, all_info)
        print(f"  数据一致性问题：{len(issues_cross_verify)} 项")
        print(f"  逻辑关系问题：{len(issues_cross_logic)} 项")

    # ════════════════════════════════════════════════════
    # Output
    # ════════════════════════════════════════════════════
    lines = []
    lines.append("=" * 72)
    lines.append("    水质检测报告验证 —— 待确认问题清单")
    lines.append("=" * 72)
    lines.append(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"扫描目录：{report_dir}")
    if orig_record_file:
        lines.append(f"原始记录：{os.path.basename(orig_record_file)}")
    lines.append(f"报告文件数：{len(files)}")
    lines.append("")

    all_sections = [
        ("一、原始记录检查", issues_original),
        ("二、交叉验证 - 数据一致性", issues_cross_verify),
        ("三、交叉验证 - 逻辑关系", issues_cross_logic),
        ("四、报告命名问题", issues_naming),
        ("五、报告编号问题", issues_numbering),
        ("六、报告数据问题", issues_data),
        ("七、报告格式/模板问题", issues_format),
        ("八、报告日期问题", issues_date),
        ("九、报告一致性问题", issues_consistency),
        ("十、报告异常值问题", issues_values),
        ("十一、文件读取问题", issues_read_errors),
    ]

    total_issues = sum(len(s[1]) for s in all_sections)
    lines.append(f"共发现 {total_issues} 项待确认问题，分类如下：")
    for title, items in all_sections:
        extra = ""
        if "异常值" in title:
            extra = f"（严重 {len(issues_values_critical)} / 一般 {len(issues_values_normal)}）"
        lines.append(f"  {title}：{len(items)} 项{extra}")
    lines.append("")

    # Build filename -> [样品XXX / 报告YYYY] tag mapping
    fname_to_tag = {}
    for fn, inf in all_info.items():
        sid = inf.get('sample_id', '')
        rn = inf.get('report_number', '')
        parts = []
        if sid:
            parts.append(f"样品{sid}")
        if rn:
            parts.append(f"报告{rn}")
        if parts:
            fname_to_tag[fn] = '[' + ' / '.join(parts) + '] '

    def _prepend_tag(issue_text):
        """Try to find a filename in the issue and prepend sample/report tag."""
        m = re.search(r'"([^"]+\.xlsx?)"', issue_text)
        if m:
            fn = m.group(1)
            tag = fname_to_tag.get(fn, '')
            if tag:
                return tag + issue_text
        return issue_text

    global_counter = 0

    def write_section(title, issues):
        nonlocal global_counter
        lines.append("-" * 72)
        lines.append(f"{title}（共 {len(issues)} 项）")
        lines.append("-" * 72)
        if not issues:
            lines.append("  （无）")
        for issue in issues:
            global_counter += 1
            tagged = _prepend_tag(issue)
            lines.append(f"  {global_counter}. {tagged}")
        lines.append("")

    for title, items in all_sections:
        write_section(title, items)

    lines.append("=" * 72)
    lines.append("以上问题均为程序自动检测，可能存在误报，请人工逐项核实。")
    lines.append("=" * 72)

    output_text = '\n'.join(lines)

    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(output_text)

    print(f"\n分析完成！共发现 {total_issues} 项待确认问题。")
    print(f"结果已写入：{output_file}")

    return output_text


if __name__ == '__main__':
    main()
