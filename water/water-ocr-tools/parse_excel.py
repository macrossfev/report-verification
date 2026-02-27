#!/usr/bin/env python3
"""
Excel parser that extracts ALL data from ALL sheets and outputs as JSON.
Usage: uv run --with openpyxl python parse_excel.py "/path/to/file.xlsx"
"""

import sys
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def parse_excel(file_path: str) -> dict:
    """
    Parse Excel file and extract all data from all sheets.

    Returns a dictionary with structure:
    {
        "sheets": {
            "SheetName": {
                "dimensions": {"rows": int, "cols": int},
                "data": [[cell_values...], ...],
                "raw_data": [[cell_values...], ...]  # Includes formulas
            }
        }
    }
    """
    workbook = load_workbook(file_path, data_only=False)
    result = {"sheets": {}}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Get sheet dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column

        # Extract all cell values
        data = []
        raw_data = []

        for row_idx in range(1, max_row + 1):
            row_values = []
            row_raw = []

            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)

                # Get computed value
                value = cell.value
                if value is not None:
                    # Convert to string if it's a date or other object
                    if hasattr(value, 'isoformat'):  # datetime objects
                        value = value.isoformat()
                    row_values.append(value)
                else:
                    row_values.append(None)

                # Get raw value (formula if present)
                raw_value = cell.value if not hasattr(cell, '_value') else cell._value
                if raw_value is not None:
                    if hasattr(raw_value, 'isoformat'):
                        raw_value = raw_value.isoformat()
                    row_raw.append(raw_value)
                else:
                    row_raw.append(None)

            data.append(row_values)
            raw_data.append(row_raw)

        result["sheets"][sheet_name] = {
            "dimensions": {
                "rows": max_row,
                "cols": max_col
            },
            "data": data,
            "raw_data": raw_data
        }

    return result


def main():
    if len(sys.argv) < 2:
        print("Usage: uv run --with openpyxl python parse_excel.py <excel_file>", file=sys.stderr)
        sys.exit(1)

    file_path = sys.argv[1]

    try:
        result = parse_excel(file_path)
        print(json.dumps(result, indent=2, ensure_ascii=False))
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
