#!/usr/bin/env python3
"""
Comprehensive analysis of ALL 23 report files (0189-0211).
Reads every sheet, every cell, prints all data.
"""

import os
import sys
import openpyxl
import xlrd
from pathlib import Path

REPORT_DIR = Path("/root/projects/report-verification/report/0189-0211")
EXCLUDE = "260202-1-24.xlsx"

SEPARATOR = "=" * 100
SUB_SEP = "-" * 80

def read_xlsx(filepath):
    """Read .xlsx file using openpyxl, return all sheets with data."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
            row_data = []
            for cell in row:
                val = cell.value
                row_data.append(val)
            rows.append(row_data)
        # Also capture merged cells info
        merged = []
        for mc in ws.merged_cells.ranges:
            merged.append(str(mc))
        result[sheet_name] = {
            'rows': rows,
            'max_row': ws.max_row,
            'max_col': ws.max_column,
            'merged_cells': merged
        }
    wb.close()
    return result


def read_xls(filepath):
    """Read .xls file using xlrd, return all sheets with data."""
    wb = xlrd.open_workbook(filepath, formatting_info=True)
    result = {}
    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        sheet_name = ws.name
        rows = []
        for row_idx in range(ws.nrows):
            row_data = []
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)
                val = cell.value
                # Convert float that looks like int
                if cell.ctype == xlrd.XL_CELL_NUMBER and val == int(val):
                    val = int(val)
                row_data.append(val)
            rows.append(row_data)
        # Get merged cells
        merged = []
        for crange in ws.merged_cells:
            merged.append(f"R{crange[0]+1}C{crange[2]+1}:R{crange[1]}C{crange[3]}")
        result[sheet_name] = {
            'rows': rows,
            'max_row': ws.nrows,
            'max_col': ws.ncols,
            'merged_cells': merged
        }
    return result


def format_cell(val):
    """Format a cell value for display."""
    if val is None:
        return "[空]"
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    return str(val)


def print_sheet_data(sheet_name, sheet_data):
    """Print all data from a sheet."""
    print(f"\n  Sheet: 【{sheet_name}】")
    print(f"  Dimensions: {sheet_data['max_row']} rows x {sheet_data['max_col']} cols")
    if sheet_data['merged_cells']:
        print(f"  Merged cells: {', '.join(sheet_data['merged_cells'][:20])}")
        if len(sheet_data['merged_cells']) > 20:
            print(f"    ... and {len(sheet_data['merged_cells'])-20} more merged ranges")
    print()

    rows = sheet_data['rows']
    if not rows:
        print("  [Empty sheet]")
        return

    # Print every row with row number
    for i, row in enumerate(rows):
        cells = [format_cell(c) for c in row]
        # Skip rows that are entirely empty
        if all(c == "[空]" for c in cells):
            continue
        print(f"  Row {i+1:3d}: | {' | '.join(cells)} |")


def extract_report_info(filename, all_sheets_data):
    """Try to extract key report metadata."""
    info = {
        'report_number': '',
        'facility': '',
        'water_type': '',
        'sample_info': '',
        'test_params': [],
    }

    # Extract from filename
    basename = os.path.splitext(filename)[0]
    # Report number is first 4 digits
    if basename[:4].isdigit():
        info['report_number'] = basename[:4]

    # Water type from filename
    if '出厂水' in filename:
        info['water_type'] = '出厂水'
    elif '原水' in filename:
        info['water_type'] = '原水'
    elif '管网水' in filename:
        info['water_type'] = '管网水'

    # Facility name
    name_part = basename[4:]
    info['facility'] = name_part

    # Try to extract from first sheet data
    for sheet_name, sdata in all_sheets_data.items():
        rows = sdata['rows']
        for row in rows:
            for cell in row:
                if cell is not None:
                    s = str(cell)
                    if '样品名称' in s or '受检单位' in s or '检测项目' in s:
                        info['sample_info'] += f" {s}"
        break  # only first sheet

    return info


def main():
    files = sorted([f for f in os.listdir(REPORT_DIR) if f != EXCLUDE and (f.endswith('.xlsx') or f.endswith('.xls'))])

    print(f"Total report files found: {len(files)}")
    print(SEPARATOR)

    # Summary tracking
    all_reports = []

    for idx, filename in enumerate(files, 1):
        filepath = REPORT_DIR / filename
        ext = os.path.splitext(filename)[1].lower()

        print(f"\n{SEPARATOR}")
        print(f"FILE {idx}/{len(files)}: {filename}")
        print(f"Path: {filepath}")
        print(f"Size: {os.path.getsize(filepath)} bytes")
        print(SEPARATOR)

        try:
            if ext == '.xlsx':
                sheets_data = read_xlsx(filepath)
            elif ext == '.xls':
                sheets_data = read_xls(filepath)
            else:
                print(f"  SKIPPED: unsupported format {ext}")
                continue
        except Exception as e:
            print(f"  ERROR reading file: {e}")
            import traceback
            traceback.print_exc()
            continue

        print(f"  Number of sheets: {len(sheets_data)}")
        print(f"  Sheet names: {list(sheets_data.keys())}")

        for sheet_name, sdata in sheets_data.items():
            print(SUB_SEP)
            print_sheet_data(sheet_name, sdata)

        # Extract summary info
        info = extract_report_info(filename, sheets_data)
        info['filename'] = filename
        all_reports.append(info)

        print(SUB_SEP)
        print(f"  >> Report #{info['report_number']} | Facility: {info['facility']} | Water Type: {info['water_type']}")
        print(SUB_SEP)

    # Final summary
    print(f"\n\n{'#' * 100}")
    print("SUMMARY OF ALL 23 REPORTS")
    print(f"{'#' * 100}")
    print(f"{'No.':<5} {'Report#':<8} {'Water Type':<10} {'Facility':<50} {'Format':<6}")
    print("-" * 85)
    for i, r in enumerate(all_reports, 1):
        ext = os.path.splitext(r['filename'])[1]
        print(f"{i:<5} {r['report_number']:<8} {r['water_type']:<10} {r['facility']:<50} {ext:<6}")

    print(f"\nTotal files processed: {len(all_reports)}")

    # Structure analysis
    print(f"\n\n{'#' * 100}")
    print("REPORT STRUCTURE ANALYSIS")
    print(f"{'#' * 100}")

    # Group by water type
    by_type = {}
    for r in all_reports:
        wt = r['water_type'] or 'unknown'
        by_type.setdefault(wt, []).append(r)

    for wt, reports in by_type.items():
        print(f"\n  Water Type: {wt} ({len(reports)} reports)")
        for r in reports:
            print(f"    - {r['report_number']} {r['facility']}")


if __name__ == "__main__":
    main()
