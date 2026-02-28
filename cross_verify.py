#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
报告163-180 vs 原始记录(260204-1-18.xlsx) 自动交叉验证
"""
import os, re
from collections import defaultdict
import openpyxl
import xlrd

BASE = "/root/projects/report-verification/report/162-188"
ORIG = os.path.join(BASE, "260204-1-18.xlsx")

# ═══ 样品编号 → 报告编号 映射 ═══
SAMPLE_MAP = {
    'W260204C01': (163, '秀山第二水厂出厂水'),
    'W260204C02': (164, '秀山第三水厂出厂水'),
    'W260204C03': (165, '秀山第二水厂管网水'),
    'W260204C04': (166, '秀山第三水厂管网水'),
    'W260204C05': (167, '秀山第二水厂原水'),
    'W260204C06': (168, '秀山第三水厂原水'),
    'W260204C07': (169, '白家湾水厂出厂水'),
    'W260204C08': (170, '三元宫水厂出厂水'),
    'W260204C09': (171, '白家湾水厂管网水'),
    'W260204C10': (172, '三元宫水厂管网水'),
    'W260204C11': (173, '三元宫水厂原水'),
    'W260204C12': (174, '白家湾水厂原水'),
    'W260204C13': (175, '环城路二水厂出厂水'),
    'W260204C14': (176, '正阳水厂出厂水'),
    'W260204C15': (177, '环城路二水厂管网水'),
    'W260204C16': (178, '正阳水厂管网水'),
    'W260204C17': (179, '小坝二级水库原水'),
    'W260204C18': (180, '正阳水厂原水'),
}

# 报告编号 → 样品编号 反向映射
RNUM_TO_SID = {rnum: sid for sid, (rnum, desc) in SAMPLE_MAP.items()}

# 被检单位映射(从Sheet1)
COMPANY_MAP = {
    'W260204C01': '重庆水务环境控股集团渝东南自来水有限公司秀山分公司',
    'W260204C02': '重庆水务环境控股集团渝东南自来水有限公司秀山分公司',
    'W260204C03': '重庆水务环境控股集团渝东南自来水有限公司秀山分公司',
    'W260204C04': '重庆水务环境控股集团渝东南自来水有限公司秀山分公司',
    'W260204C05': '重庆水务环境控股集团渝东南自来水有限公司秀山分公司',
    'W260204C06': '重庆水务环境控股集团渝东南自来水有限公司秀山分公司',
    'W260204C07': '重庆水务环境控股集团渝东南自来水有限公司',
    'W260204C08': '重庆水务环境控股集团渝东南自来水有限公司',
    'W260204C09': '重庆水务环境控股集团渝东南自来水有限公司',
    'W260204C10': '重庆水务环境控股集团渝东南自来水有限公司',
    'W260204C11': '重庆水务环境控股集团渝东南自来水有限公司',
    'W260204C12': '重庆水务环境控股集团渝东南自来水有限公司',
    'W260204C13': '重庆水务环境控股集团渝东南自来水有限公司酉阳分公司',
    'W260204C14': '重庆水务环境控股集团渝东南自来水有限公司',
    'W260204C15': '重庆水务环境控股集团渝东南自来水有限公司酉阳分公司',
    'W260204C16': '重庆水务环境控股集团渝东南自来水有限公司',
    'W260204C17': '重庆水务环境控股集团渝东南自来水有限公司酉阳分公司',
    'W260204C18': '重庆水务环境控股集团渝东南自来水有限公司',
}


def clean_item_name(raw):
    """清理检测项目名称（处理多行表头等情况）"""
    s = str(raw).strip()
    s = re.sub(r'\s*\n\s*', '', s)  # join multi-line
    s = re.sub(r'\s*[\(（][^)）]*[\)）]\s*$', '', s).strip()  # remove trailing unit
    s = re.sub(r'\s*[\(（][^)）]*$', '', s).strip()  # handle unclosed parens
    s = re.sub(r'\s{2,}', ' ', s).strip()
    return s


def normalize_val(v):
    """标准化数值用于比较"""
    if v is None:
        return None
    s = str(v).strip()
    s = s.replace('＜', '<').replace('≤', '<=')
    # Try to parse as float for comparison
    try:
        return float(s)
    except:
        return s


def vals_match(orig_val, report_val):
    """比较原始记录值与报告值是否一致"""
    if orig_val is None or report_val is None:
        return True  # skip if either is missing

    o = str(orig_val).strip().replace('＜', '<')
    r = str(report_val).strip().replace('＜', '<')

    if o == r:
        return True

    # Try numeric comparison
    try:
        of = float(o.replace('<', ''))
        rf = float(r.replace('<', ''))
        if abs(of - rf) < 0.0001:
            return True
    except:
        pass

    # Handle "0" vs "未检出" for bacteria
    if (o == '0' and r in ('未检出', '0')) or (r == '0' and o in ('未检出', '0')):
        return True

    return False


def read_original_data():
    """从原始记录读取所有样品的检测数据"""
    wb = openpyxl.load_workbook(ORIG, data_only=True)
    data = defaultdict(dict)  # sample_id -> {item_name: value}
    # 通用样品编号模式：W/K/M + 6位日期 + C + 数字
    sid_pattern = r'[WKM]\d{6}C\d+'

    for sname in wb.sheetnames:
        if sname == 'Sheet1':
            continue
        ws = wb[sname]

        # Find sample IDs in header row (row 2 or 3) with dynamic start row
        sample_cols = {}  # col_index -> sample_id
        header_row_max = 0
        for hr in [2, 3]:
            for c in range(1, ws.max_column + 1):
                v = ws.cell(hr, c).value
                if v and re.match(sid_pattern, str(v).strip()):
                    sample_cols[c] = str(v).strip()
                    header_row_max = max(header_row_max, hr)

        if not sample_cols:
            # Try different layout (sample IDs in column 1)
            for r in range(1, ws.max_row + 1):
                v = ws.cell(r, 1).value
                if v and re.match(sid_pattern, str(v).strip()):
                    sid = str(v).strip()
                    for c in range(2, ws.max_column + 1):
                        header = ws.cell(3, c).value
                        val = ws.cell(r, c).value
                        if header and val is not None:
                            item_name = clean_item_name(header)
                            if item_name:
                                data[sid][item_name] = str(val).strip()
            continue

        # Dynamic data start row based on header position
        data_start = header_row_max + 1

        # Normal layout: items in rows, samples in columns
        for r in range(data_start, ws.max_row + 1):
            item_cell = ws.cell(r, 1).value
            if not item_cell:
                continue
            cname = clean_item_name(item_cell)
            if not cname:
                continue

            for c, sid in sample_cols.items():
                val = ws.cell(r, c).value
                if val is not None and str(val).strip():
                    data[sid][cname] = str(val).strip()

    wb.close()
    return dict(data)


def read_report(filepath):
    """读取报告文件"""
    fname = os.path.basename(filepath)
    info = {}

    if fname.endswith('.xlsx'):
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws1 = wb[wb.sheetnames[0]]
        b1 = ws1.cell(1, 2).value
        if b1:
            m = re.search(r'第\s*\(\s*(\d+)\s*\)\s*号', str(b1))
            if m:
                info['report_number'] = int(m.group(1))

        for r in range(7, 13):
            cv = ws1.cell(r, 3).value
            if cv and ('水' in str(cv) or '【' in str(cv)):
                info['sample_name'] = str(cv).strip()
                break
        for r in range(8, 13):
            cv = ws1.cell(r, 3).value
            if cv and '公司' in str(cv):
                info['company'] = str(cv).strip()
                break
        for r in range(10, 14):
            bv = ws1.cell(r, 2).value
            cv = ws1.cell(r, 3).value
            if bv and '报告编制日期' in str(bv) and cv:
                info['report_date'] = str(cv).strip()
                break

        if len(wb.sheetnames) >= 2:
            ws2 = wb[wb.sheetnames[1]]
            info['sample_id'] = str(ws2.cell(8, 3).value or '').strip()
            info['sample_type'] = str(ws2.cell(3, 3).value or '').strip()
            info['sampling_date'] = str(ws2.cell(4, 5).value or '').strip()
            info['receipt_date'] = str(ws2.cell(5, 5).value or '').strip()
            info['testing_date'] = str(ws2.cell(8, 5).value or '').strip()
            info['test_items_desc'] = str(ws2.cell(10, 3).value or '').strip()
            info['conclusion'] = str(ws2.cell(13, 2).value or '').strip()

        test_items = []
        for si in range(2, len(wb.sheetnames)):
            ws = wb[wb.sheetnames[si]]
            for r in range(1, ws.max_row + 1):
                a = ws.cell(r, 1).value
                b = ws.cell(r, 2).value
                d = ws.cell(r, 4).value
                if a is not None and b is not None:
                    try:
                        seq = int(float(str(a)))
                        if 1 <= seq <= 100:
                            test_items.append({
                                'seq': seq,
                                'name': str(b).strip(),
                                'result': str(d).strip() if d is not None else '',
                                'standard': str(ws.cell(r, 5).value or '').strip(),
                            })
                    except:
                        pass
        info['test_items'] = test_items
        wb.close()
    else:
        wb = xlrd.open_workbook(filepath)
        ws1 = wb.sheet_by_index(0)
        if ws1.nrows > 0 and ws1.ncols > 1:
            b1 = ws1.cell_value(0, 1)
            if b1:
                m = re.search(r'第\s*\(\s*(\d+)\s*\)\s*号', str(b1))
                if m:
                    info['report_number'] = int(m.group(1))
        for r in range(6, min(12, ws1.nrows)):
            if ws1.ncols > 2:
                cv = ws1.cell_value(r, 2)
                if cv and ('水' in str(cv) or '【' in str(cv)):
                    info['sample_name'] = str(cv).strip()
                    break
        for r in range(7, min(12, ws1.nrows)):
            if ws1.ncols > 2:
                cv = ws1.cell_value(r, 2)
                if cv and '公司' in str(cv):
                    info['company'] = str(cv).strip()
                    break
        for r in range(9, min(14, ws1.nrows)):
            if ws1.ncols > 2:
                bv = ws1.cell_value(r, 1)
                cv = ws1.cell_value(r, 2)
                if bv and '报告编制日期' in str(bv) and cv:
                    info['report_date'] = str(cv).strip()
                    break

        if wb.nsheets >= 2:
            ws2 = wb.sheet_by_index(1)
            def sv(r, c):
                if r < ws2.nrows and c < ws2.ncols:
                    return ws2.cell_value(r, c)
                return None
            info['sample_id'] = str(sv(7, 2) or '').strip()
            info['sample_type'] = str(sv(2, 2) or '').strip()
            info['sampling_date'] = str(sv(3, 4) or '').strip()
            info['receipt_date'] = str(sv(4, 4) or '').strip()
            info['testing_date'] = str(sv(7, 4) or '').strip()
            info['test_items_desc'] = str(sv(9, 2) or '').strip()
            info['conclusion'] = str(sv(12, 1) or '').strip()

        test_items = []
        for si in range(2, wb.nsheets):
            ws = wb.sheet_by_index(si)
            for r in range(ws.nrows):
                if ws.ncols >= 6:
                    a = ws.cell_value(r, 0)
                    b = ws.cell_value(r, 1)
                    d = ws.cell_value(r, 3)
                    if a not in ('', None) and b not in ('', None):
                        try:
                            seq = int(float(str(a)))
                            if 1 <= seq <= 100:
                                test_items.append({
                                    'seq': seq,
                                    'name': str(b).strip(),
                                    'result': str(d).strip() if d not in ('', None) else '',
                                    'standard': str(ws.cell_value(r, 4) or '').strip(),
                                })
                        except:
                            pass
        info['test_items'] = test_items

    return info


# ═══ 名称匹配映射(原始记录项目名 → 报告项目名) ═══
NAME_ALIAS = {
    '游离氯': '游离氯',
    '二氧化氯': '二氧化氯',
    '水温': '水温',
    'pH': 'pH',
    '色度': '色度',
    '浑浊度': '浑浊度',
    '高锰酸盐指数': '高锰酸盐指数(以O2计)',
    '溶解氧': '溶解氧',
    '化学需氧量': '化学需氧量(COD)',
    '五日生化需氧量': '五日生化需氧量(BOD5)',
    '菌落总数': '菌落总数',
    '总大肠菌群': '总大肠菌群',
    '大肠埃希氏菌': '大肠埃希氏菌',
    '粪大肠菌群': '粪大肠菌群',
    '电导率': '电导率',
    '氟化物': '氟化物',
    '氯化物': '氯化物',
    '硝酸盐(以N计)': '硝酸盐(以N计)',
    '硫酸盐': '硫酸盐',
    '亚氯酸盐': '亚氯酸盐',
    '氯酸盐': '氯酸盐',
    '二氯乙酸': '二氯乙酸',
    '三氯乙酸': '三氯乙酸',
    '铜': '铜',
    '铁': '铁',
    '锰': '锰',
    '砷': '砷',
    '锌': '锌',
    '硒': '硒',
    '汞': '汞',
    '镉': '镉',
    '铅': '铅',
    '铝': '铝',
    '阴离子合成洗涤剂': '阴离子合成洗涤剂',
    '阴离子表面活性剂': '阴离子表面活性剂',
    '挥发酚': '挥发酚类(以苯酚计)',
    '氰化物': '氰化物',
    '六 价 铬': '铬(六价)',
    '六价铬': '铬(六价)',
    '总硬度(以': '总硬度(以CaCO3计)',
    '氨(以N计)': '氨(以N计)',
    '溶解性总固体': '溶解性总固体',
    '三氯甲烷': '三氯甲烷',
    '四氯化碳': '四氯化碳',
    '二氯一溴甲烷': '二氯一溴甲烷',
    '一氯二溴甲烷': '一氯二溴甲烷',
    '三溴甲烷': '三溴甲烷',
    '三卤甲烷': '三卤甲烷',
    '总a': '总α放射性',
    '总β': '总β放射性',
    '总磷': '总磷(以P计)',
    '总氮': '总氮(以N计)',
    '氨氮': '氨氮(NH3-N)',
    '硫化物': '硫化物',
    '石油类': '石油类',
    '钙': '钙',
    '镁': '镁',
}


def find_report_item(test_items, orig_name):
    """在报告的检测项目中查找匹配的项目"""
    # Direct match
    for item in test_items:
        if item['name'] == orig_name:
            return item

    # Alias match
    report_name = NAME_ALIAS.get(orig_name)
    if report_name:
        for item in test_items:
            if item['name'] == report_name:
                return item

    # Fuzzy match
    for item in test_items:
        # Check if one contains the other
        if orig_name in item['name'] or item['name'] in orig_name:
            return item
        # Try without parenthetical suffixes
        clean_orig = re.sub(r'\(.*?\)', '', orig_name).strip()
        clean_item = re.sub(r'\(.*?\)', '', item['name']).strip()
        if clean_orig and clean_item and (clean_orig in clean_item or clean_item in clean_orig):
            return item

    return None


def check_original_records(orig_data, sample_map, plant_groups_orig=None):
    """原始记录自查：检查原始记录自身的逻辑一致性（改进#1）

    Args:
        orig_data: dict of sample_id -> {item_name: value}
        sample_map: dict of sample_id -> (report_num, description, ...)
        plant_groups_orig: optional dict of plant_name -> {water_type: sample_id}
            如果为None，则自动从sample_map推断
    Returns:
        list of (severity, report_num, sample_ids, description) issues
    """
    issues = []

    # Auto-detect plant groups from sample_map if not provided
    if plant_groups_orig is None:
        plant_groups_orig = defaultdict(dict)
        for sid, info in sample_map.items():
            if not sid.startswith('W'):
                continue
            desc = info[1] if len(info) > 1 else ''
            # Extract plant name and water type from description
            for wtype in ['出厂水', '管网末梢水', '管网水', '原水']:
                if wtype in desc:
                    plant_name = desc.replace(wtype, '').strip()
                    plant_groups_orig[plant_name][wtype] = sid
                    break

    # 1. 同一水厂出厂水/管网水/原水间逻辑关系
    for plant, sids in plant_groups_orig.items():
        cc_sid = sids.get('出厂水')
        gw_sid = sids.get('管网水') or sids.get('管网末梢水')
        yw_sid = sids.get('原水')

        # 游离氯：管网应<=出厂
        if cc_sid and gw_sid:
            cc_data = orig_data.get(cc_sid, {})
            gw_data = orig_data.get(gw_sid, {})
            try:
                cc_cl = float(cc_data.get('游离氯', '0'))
                gw_cl = float(gw_data.get('游离氯', '0'))
                if gw_cl > cc_cl * 1.1 and gw_cl > 0 and cc_cl > 0:
                    issues.append(('注意-原始记录', 0, f"{gw_sid},{cc_sid}",
                        f"[原始记录] {plant}管网水游离氯({gw_cl})高于出厂水({cc_cl})，需确认"))
            except (ValueError, TypeError):
                pass

        # 高锰酸盐指数：出厂水应<=原水
        if cc_sid and yw_sid:
            cc_data = orig_data.get(cc_sid, {})
            yw_data = orig_data.get(yw_sid, {})
            try:
                cc_v = cc_data.get('高锰酸盐指数', '')
                yw_v = yw_data.get('高锰酸盐指数', '')
                if cc_v and yw_v:
                    cc_f = float(str(cc_v).replace('<', ''))
                    yw_f = float(str(yw_v).replace('<', ''))
                    if cc_f > yw_f * 1.2:
                        issues.append(('注意-原始记录', 0, f"{cc_sid},{yw_sid}",
                            f"[原始记录] {plant}出厂水高锰酸盐指数({cc_f})高于原水({yw_f})，异常"))
            except (ValueError, TypeError):
                pass

    # 2. pH范围检查
    for sid, items in orig_data.items():
        if not sid.startswith('W'):
            continue
        info = sample_map.get(sid)
        label = f"{info[1]}({sid})" if info else sid
        ph_val = items.get('pH')
        if ph_val:
            try:
                ph = float(str(ph_val).replace('<', ''))
                if ph < 5 or ph > 10:
                    issues.append(('严重-原始记录', 0, sid,
                        f"[原始记录] {label} pH值={ph}异常（通常范围5-10）"))
            except (ValueError, TypeError):
                pass

    # 3. 质控样品(M/K系列)异常值检查
    for sid, items in orig_data.items():
        if not (sid.startswith('M') or sid.startswith('K')):
            continue
        for item_name, val in items.items():
            try:
                v = float(str(val).replace('<', ''))
                if v < 0:
                    issues.append(('严重-原始记录', 0, sid,
                        f"[原始记录] 质控样品{sid}项目「{item_name}」值为负数({val})"))
            except (ValueError, TypeError):
                pass

    # 4. 同水源不同水厂数据一致性检查
    # 找出所有原水样品，按描述中的水源名分组
    source_groups = defaultdict(list)
    for sid, info in sample_map.items():
        if not sid.startswith('W'):
            continue
        desc = info[1] if len(info) > 1 else ''
        if '原水' in desc:
            source_groups[desc].append(sid)
    # 同一水源的不同原水样品数据应相近（这里检查有多个样品的水源）
    for source_desc, sids_list in source_groups.items():
        if len(sids_list) < 2:
            continue
        for param in ['pH', '高锰酸盐指数', '溶解氧']:
            vals = {}
            for sid in sids_list:
                v = orig_data.get(sid, {}).get(param)
                if v:
                    try:
                        vals[sid] = float(str(v).replace('<', ''))
                    except (ValueError, TypeError):
                        pass
            if len(vals) >= 2:
                vlist = list(vals.values())
                if min(vlist) > 0 and max(vlist) / min(vlist) > 2:
                    detail = ', '.join(f"{s}={v}" for s, v in vals.items())
                    issues.append(('注意-原始记录', 0, ','.join(vals.keys()),
                        f"[原始记录] 同水源「{source_desc}」各样品{param}差异较大: {detail}"))

    return issues


def main():
    issues = []  # (severity, report_num, description)

    # 1. Read original records
    orig_data = read_original_data()

    # 1b. 原始记录自查（改进#1）
    issues.extend(check_original_records(orig_data, SAMPLE_MAP))

    # 2. Read all reports
    reports = {}
    for fname in sorted(os.listdir(BASE)):
        if not fname.startswith('0') or not fname.endswith(('.xlsx', '.xls')):
            continue
        m = re.match(r'^(\d+)', fname)
        if not m:
            continue
        num = int(m.group(1))
        if num < 163 or num > 180:
            continue
        reports[num] = read_report(os.path.join(BASE, fname))
        reports[num]['filename'] = fname

    # ═══════ 验证 ═══════

    # A. 被检单位一致性检查
    for sid, (rnum, desc) in SAMPLE_MAP.items():
        if rnum not in reports:
            issues.append(('严重', rnum, f"缺少报告文件（样品{sid}，{desc}）"))
            continue
        rpt = reports[rnum]
        expected_company = COMPANY_MAP.get(sid, '')
        actual_company = rpt.get('company', '')
        if actual_company and expected_company:
            if expected_company not in actual_company and actual_company not in expected_company:
                issues.append(('重要', rnum,
                    f"被检单位不一致: 原始记录为「{expected_company}」, "
                    f"报告为「{actual_company}」"))
            elif expected_company != actual_company:
                # Partial match - might be abbreviated
                if '秀山分公司' in expected_company and '秀山' not in actual_company:
                    issues.append(('注意', rnum,
                        f"被检单位可能不完整: 原始记录为「{expected_company}」, "
                        f"报告为「{actual_company}」(缺少'秀山分公司')"))
                elif '酉阳分公司' in expected_company and '酉阳' not in actual_company:
                    issues.append(('注意', rnum,
                        f"被检单位可能不完整: 原始记录为「{expected_company}」, "
                        f"报告为「{actual_company}」(缺少'酉阳分公司')"))

    # B. 样品编号一致性
    for sid, (rnum, desc) in SAMPLE_MAP.items():
        if rnum not in reports:
            continue
        rpt = reports[rnum]
        actual_sid = rpt.get('sample_id', '')
        if actual_sid and actual_sid != sid:
            issues.append(('严重', rnum,
                f"样品编号不一致: 应为「{sid}」, 报告为「{actual_sid}」"))

    # C. 检测数据交叉比对
    for sid, (rnum, desc) in SAMPLE_MAP.items():
        if rnum not in reports or sid not in orig_data:
            continue
        rpt = reports[rnum]
        orig_items = orig_data[sid]
        test_items = rpt.get('test_items', [])

        is_raw_water = '原水' in desc

        for orig_name, orig_val in orig_items.items():
            # Skip informational items
            if orig_name in ('钙', '镁', '电导率', '水温'):
                continue
            # 改进#4: 原水报告中肉眼可见物/臭和味可能不列出
            if is_raw_water and orig_name in ('肉眼可见物', '臭和味'):
                continue

            matched_item = find_report_item(test_items, orig_name)
            if matched_item is None:
                issues.append(('注意', rnum,
                    f"原始记录中的项目「{orig_name}」(值={orig_val}) 未在报告中找到对应项"))
                continue

            report_val = matched_item['result']
            if not vals_match(orig_val, report_val):
                issues.append(('严重', rnum,
                    f"数据不一致 - 项目「{orig_name}」: "
                    f"原始记录={orig_val}, 报告={report_val} "
                    f"(报告项名: {matched_item['name']})"))

    # D. 报告间逻辑一致性检查
    # D1. 同一水厂的出厂水和管网水：消毒剂余量，管网应<=出厂
    plant_groups = {
        '秀山第二水厂': {'出厂水': 163, '管网水': 165, '原水': 167},
        '秀山第三水厂': {'出厂水': 164, '管网水': 166, '原水': 168},
        '白家湾水厂': {'出厂水': 169, '管网水': 171, '原水': 174},
        '三元宫水厂': {'出厂水': 170, '管网水': 172, '原水': 173},
        '环城路二水厂': {'出厂水': 175, '管网水': 177},
        '正阳水厂': {'出厂水': 176, '管网水': 178, '原水': 180},
    }

    def get_result_float(rnum, item_name):
        if rnum not in reports:
            return None
        for item in reports[rnum].get('test_items', []):
            if item_name in item['name']:
                v = item['result'].replace('＜', '<').replace('<', '')
                try:
                    return float(v)
                except:
                    return None
        return None

    for plant, type_map in plant_groups.items():
        if '出厂水' in type_map and '管网水' in type_map:
            ccr = type_map['出厂水']
            gwsr = type_map['管网水']

            # Check chlorine: 管网 should generally <= 出厂
            for chlorine_name in ['游离氯', '二氧化氯']:
                ccv = get_result_float(ccr, chlorine_name)
                gwv = get_result_float(gwsr, chlorine_name)
                if ccv is not None and gwv is not None:
                    if gwv > ccv * 1.1:  # Allow 10% tolerance
                        issues.append(('注意', gwsr,
                            f"{plant}管网水{chlorine_name}({gwv})高于出厂水({ccv})，需确认"))

            # 浑浊度: 出厂水应低于原水
            if '原水' in type_map:
                ywr = type_map['原水']
                cc_turb = get_result_float(ccr, '浑浊度')
                yw_turb = get_result_float(ywr, '浑浊度')
                # Don't check this - turbidity in raw water may not be directly comparable

            # 高锰酸盐指数: 出厂水应<=原水
            if '原水' in type_map:
                ywr = type_map['原水']
                cc_kmno4 = get_result_float(ccr, '高锰酸盐指数')
                yw_kmno4 = get_result_float(ywr, '高锰酸盐指数')
                if cc_kmno4 is not None and yw_kmno4 is not None:
                    if cc_kmno4 > yw_kmno4 * 1.5:
                        issues.append(('注意', ccr,
                            f"{plant}出厂水高锰酸盐指数({cc_kmno4})显著高于原水({yw_kmno4})，异常"))

    # E. 检测项目数一致性
    for rnum, rpt in reports.items():
        items_desc = rpt.get('test_items_desc', '')
        m = re.search(r'(\d+)\s*项', items_desc)
        if m:
            declared = int(m.group(1))
            actual = len(rpt.get('test_items', []))
            if declared != actual:
                issues.append(('重要', rnum,
                    f"检测项目数不一致: 声称{declared}项, 实际提取{actual}项"))

    # F. 同批次出厂水/管网水项目数应一致
    ccr_counts = {}
    gwr_counts = {}
    for rnum, rpt in reports.items():
        fname = rpt.get('filename', '')
        count = len(rpt.get('test_items', []))
        if '出厂水' in fname:
            ccr_counts[rnum] = count
        elif '管网水' in fname:
            gwr_counts[rnum] = count

    if ccr_counts:
        vals = list(ccr_counts.values())
        if len(set(vals)) > 1:
            issues.append(('注意', 0,
                f"出厂水报告检测项目数不统一: {ccr_counts}"))
    if gwr_counts:
        vals = list(gwr_counts.values())
        if len(set(vals)) > 1:
            issues.append(('注意', 0,
                f"管网水报告检测项目数不统一: {gwr_counts}"))

    # G. 超标检查
    for rnum, rpt in reports.items():
        for item in rpt.get('test_items', []):
            result = item['result'].replace('＜', '<')
            standard = item['standard']
            name = item['name']

            if not result or result.startswith('<') or result in ('无', '未检出', '无异臭、异味', '0'):
                continue
            if '水温' in name:
                continue

            try:
                val = float(result)
            except:
                continue

            std_limit = None
            # Parse standard
            std_match = re.search(r'[≤<]\s*([\d.]+)', standard)
            if std_match:
                try:
                    std_limit = float(std_match.group(1))
                except:
                    pass
            elif re.match(r'^[\d.]+$', standard):
                try:
                    std_limit = float(standard)
                except:
                    pass

            range_match = re.match(r'([\d.]+)\s*[~\-～]\s*([\d.]+)', standard)
            if range_match:
                try:
                    lo = float(range_match.group(1))
                    hi = float(range_match.group(2))
                    if val < lo:
                        issues.append(('严重', rnum,
                            f"项目「{name}」结果{result}低于标准下限{lo} (标准: {standard})"))
                    elif val > hi:
                        issues.append(('严重', rnum,
                            f"项目「{name}」结果{result}超出标准上限{hi} (标准: {standard})"))
                except:
                    pass
            elif std_limit is not None and val > std_limit:
                ratio = val / std_limit if std_limit > 0 else 0
                sev = '严重' if ratio >= 2.0 or name in ('铅', '汞', '镉', '砷') else '重要'
                issues.append((sev, rnum,
                    f"项目「{name}」结果{result}超出标准限值{std_limit} "
                    f"(标准: {standard}, {ratio:.1f}倍)"))

    # H. 报告日期格式一致性
    date_formats = {}
    for rnum, rpt in reports.items():
        rd = rpt.get('report_date', '')
        if rd:
            date_formats[rnum] = rd

    if date_formats:
        unique_formats = set()
        for rnum, d in date_formats.items():
            # Normalize to detect format differences
            fmt = re.sub(r'\d', 'D', d)
            unique_formats.add(fmt)
        if len(unique_formats) > 1:
            issues.append(('注意', 0,
                f"报告编制日期格式不统一: {date_formats}"))

    # I. 原水报告结论检查（改进#4）- 原水应引用地表水标准而非生活饮用水标准
    for rnum, rpt in reports.items():
        fname = rpt.get('filename', '')
        conclusion = rpt.get('conclusion', '')
        if '原水' in fname:
            if '符合' in conclusion and '生活饮用水' in conclusion:
                issues.append(('注意', rnum,
                    f"原水报告引用了生活饮用水标准，通常应引用地表水标准"))
            elif '符合' in conclusion:
                issues.append(('注意', rnum,
                    f"原水报告出现合格判定结论: 「{conclusion[:50]}」，请确认引用标准是否正确"))

    # J. 跨批次样品识别（改进#5）- 识别报告中样品编号不在当前原始记录中的情况
    known_sids = set(orig_data.keys())
    for rnum, rpt in reports.items():
        report_sid = rpt.get('sample_id', '')
        if report_sid and report_sid not in known_sids:
            issues.append(('注意', rnum,
                f"报告样品编号「{report_sid}」在当前原始记录文件中无对应数据，"
                f"可能属于其他批次"))

    # ═══════ 输出 ═══════
    print("=" * 80)
    print("  交叉验证结果")
    print("=" * 80)
    print()

    # Sort: 严重-原始记录 > 严重 > 重要 > 注意-原始记录 > 注意
    sev_order = {'严重-原始记录': 0, '严重': 1, '重要': 2, '注意-原始记录': 3, '注意': 4}
    issues.sort(key=lambda x: (sev_order.get(x[0], 9), x[1]))

    by_severity = defaultdict(list)
    for issue in issues:
        if len(issue) == 4:
            sev, rnum, sample_ids, desc = issue
        else:
            sev, rnum, desc = issue
            sample_ids = ''
        by_severity[sev].append((rnum, sample_ids, desc))

    total = len(issues)
    print(f"共发现 {total} 项待确认问题:")
    for sev in ['严重-原始记录', '严重', '重要', '注意-原始记录', '注意']:
        if by_severity[sev]:
            print(f"  [{sev}] {len(by_severity[sev])} 项")
    print()

    counter = 0
    for sev in ['严重-原始记录', '严重', '重要', '注意-原始记录', '注意']:
        items = by_severity[sev]
        if not items:
            continue
        print(f"\n{'─' * 80}")
        print(f"  [{sev}] 共 {len(items)} 项")
        print(f"{'─' * 80}")
        for rnum, sample_ids, desc in items:
            counter += 1
            if sample_ids:
                rpt_tag = f"样品{sample_ids}"
                if rnum > 0:
                    rpt_tag += f" / 报告{rnum:04d}"
            elif rnum > 0:
                sid = RNUM_TO_SID.get(rnum, '')
                rpt_tag = f"样品{sid} / 报告{rnum:04d}" if sid else f"报告{rnum:04d}"
            else:
                rpt_tag = "全局"
            print(f"  {counter:2d}. [{rpt_tag}] {desc}")

    print(f"\n{'═' * 80}")
    print("以上问题由程序自动检测，可能存在误报，请人工逐项核实。")
    print(f"{'═' * 80}")


if __name__ == '__main__':
    main()
