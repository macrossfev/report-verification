#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
报告163-180与原始记录(260204-1-18.xlsx)交叉验证脚本
"""

import os, re, sys, json
from collections import defaultdict
from datetime import datetime

import openpyxl
import xlrd

BASE_DIR = "/root/projects/report-verification/report/162-188"
ORIGINAL_FILE = os.path.join(BASE_DIR, "260204-1-18.xlsx")

# ════════════════════════════════════════
# 1. 读取原始记录
# ════════════════════════════════════════

def read_original_records():
    """读取260204-1-18.xlsx中的所有原始记录数据"""
    wb = openpyxl.load_workbook(ORIGINAL_FILE, data_only=True)

    result = {
        'sheets': {},
        'sample_registry': [],
        'test_data': {}  # sample_id -> {item_name: value}
    }

    print(f"原始记录文件: {ORIGINAL_FILE}")
    print(f"工作表: {wb.sheetnames}")
    print()

    # Sheet1: 样品登记表
    ws1 = wb[wb.sheetnames[0]]
    print(f"=== Sheet1 (样品登记表) ===")
    print(f"行数: {ws1.max_row}, 列数: {ws1.max_column}")

    # Dump all content of Sheet1
    print("\n--- Sheet1 全部内容 ---")
    for r in range(1, ws1.max_row + 1):
        row_data = []
        for c in range(1, min(ws1.max_column + 1, 15)):
            v = ws1.cell(r, c).value
            if v is not None:
                row_data.append(f"[{c}]{v}")
        if row_data:
            print(f"  行{r}: {' | '.join(row_data)}")

    # Read remaining sheets (test data)
    for si in range(1, len(wb.sheetnames)):
        sname = wb.sheetnames[si]
        ws = wb[sname]
        print(f"\n=== {sname} ===")
        print(f"行数: {ws.max_row}, 列数: {ws.max_column}")

        # Dump all content
        for r in range(1, ws.max_row + 1):
            row_data = []
            for c in range(1, min(ws.max_column + 1, 20)):
                v = ws.cell(r, c).value
                if v is not None:
                    row_data.append(f"[{c}]{v}")
            if row_data:
                print(f"  行{r}: {' | '.join(row_data)}")

    wb.close()
    return result


# ════════════════════════════════════════
# 2. 读取报告文件
# ════════════════════════════════════════

def read_xlsx_full(filepath):
    """完整读取xlsx报告"""
    info = {'filename': os.path.basename(filepath)}
    wb = openpyxl.load_workbook(filepath, data_only=True)
    info['sheet_names'] = wb.sheetnames
    info['sheets_data'] = {}

    for sname in wb.sheetnames:
        ws = wb[sname]
        sheet_data = []
        for r in range(1, ws.max_row + 1):
            row = {}
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v is not None:
                    row[c] = v
            if row:
                sheet_data.append((r, row))
        info['sheets_data'][sname] = sheet_data

    # Extract test items from sheet 3+
    test_items = []
    for si in range(2, len(wb.sheetnames)):
        ws = wb[wb.sheetnames[si]]
        for r in range(1, ws.max_row + 1):
            a = ws.cell(r, 1).value
            b = ws.cell(r, 2).value
            c_val = ws.cell(r, 3).value
            d = ws.cell(r, 4).value
            e = ws.cell(r, 5).value
            f = ws.cell(r, 6).value
            if a is not None and b is not None:
                try:
                    seq = int(float(str(a)))
                    if 1 <= seq <= 100:
                        test_items.append({
                            'seq': seq,
                            'name': str(b).strip(),
                            'unit': str(c_val or '').strip(),
                            'result': str(d).strip() if d is not None else '',
                            'standard': str(e or '').strip(),
                            'method': str(f or '').strip(),
                        })
                except (ValueError, TypeError):
                    pass
    info['test_items'] = test_items

    # Extract metadata from sheet2
    if len(wb.sheetnames) >= 2:
        ws2 = wb[wb.sheetnames[1]]
        info['sample_type'] = str(ws2.cell(3, 3).value or '').strip()
        info['sampler'] = str(ws2.cell(4, 3).value or '').strip()
        info['sampling_date'] = str(ws2.cell(4, 5).value or '').strip()
        info['receipt_date'] = str(ws2.cell(5, 5).value or '').strip()
        info['sampling_location'] = str(ws2.cell(6, 3).value or '').strip()
        info['sample_id'] = str(ws2.cell(8, 3).value or '').strip()
        info['testing_date'] = str(ws2.cell(8, 5).value or '').strip()
        info['product_standard'] = str(ws2.cell(9, 3).value or '').strip()
        info['test_items_desc'] = str(ws2.cell(10, 3).value or '').strip()
        info['conclusion'] = str(ws2.cell(13, 2).value or '').strip()

        # Report number from sheet1
        ws1 = wb[wb.sheetnames[0]]
        b1 = ws1.cell(1, 2).value
        if b1:
            info['report_number_raw'] = str(b1).strip()
            m = re.search(r'第\s*\(\s*(\d+)\s*\)\s*号', str(b1))
            if m:
                info['report_number'] = m.group(1).strip()

        # Report date
        for r in range(10, 14):
            bv = ws1.cell(r, 2).value
            cv = ws1.cell(r, 3).value
            if bv and '报告编制日期' in str(bv) and cv:
                info['report_date'] = str(cv).strip()
                break

        # Sample name
        for r in range(7, 13):
            cv = ws1.cell(r, 3).value
            if cv and ('水' in str(cv) or '【' in str(cv)):
                info['sample_name'] = str(cv).strip()
                break

        # Company
        for r in range(8, 13):
            cv = ws1.cell(r, 3).value
            if cv and '公司' in str(cv):
                info['company'] = str(cv).strip()
                break

    wb.close()
    return info


def read_xls_full(filepath):
    """完整读取xls报告"""
    info = {'filename': os.path.basename(filepath)}
    wb = xlrd.open_workbook(filepath)
    info['sheet_names'] = wb.sheet_names()
    info['sheets_data'] = {}

    for si in range(wb.nsheets):
        sname = wb.sheet_names()[si]
        ws = wb.sheet_by_index(si)
        sheet_data = []
        for r in range(ws.nrows):
            row = {}
            for c in range(ws.ncols):
                v = ws.cell_value(r, c)
                if v not in ('', None):
                    row[c + 1] = v  # 1-indexed to match xlsx
            if row:
                sheet_data.append((r + 1, row))  # 1-indexed
        info['sheets_data'][sname] = sheet_data

    # Extract test items
    test_items = []
    for si in range(2, wb.nsheets):
        ws = wb.sheet_by_index(si)
        for r in range(ws.nrows):
            if ws.ncols >= 6:
                a = ws.cell_value(r, 0)
                b = ws.cell_value(r, 1)
                c_val = ws.cell_value(r, 2)
                d = ws.cell_value(r, 3)
                e = ws.cell_value(r, 4)
                f = ws.cell_value(r, 5)
                if a not in ('', None) and b not in ('', None):
                    try:
                        seq = int(float(str(a)))
                        if 1 <= seq <= 100:
                            test_items.append({
                                'seq': seq,
                                'name': str(b).strip(),
                                'unit': str(c_val or '').strip(),
                                'result': str(d).strip() if d not in ('', None) else '',
                                'standard': str(e or '').strip(),
                                'method': str(f or '').strip(),
                            })
                    except (ValueError, TypeError):
                        pass
    info['test_items'] = test_items

    # Metadata
    if wb.nsheets >= 2:
        ws2 = wb.sheet_by_index(1)
        def sv(r, c):
            if r < ws2.nrows and c < ws2.ncols:
                return ws2.cell_value(r, c)
            return None

        info['sample_type'] = str(sv(2, 2) or '').strip()
        info['sampler'] = str(sv(3, 2) or '').strip()
        info['sampling_date'] = str(sv(3, 4) or '').strip()
        info['receipt_date'] = str(sv(4, 4) or '').strip()
        info['sampling_location'] = str(sv(5, 2) or '').strip()
        info['sample_id'] = str(sv(7, 2) or '').strip()
        info['testing_date'] = str(sv(7, 4) or '').strip()
        info['product_standard'] = str(sv(8, 2) or '').strip()
        info['test_items_desc'] = str(sv(9, 2) or '').strip()
        info['conclusion'] = str(sv(12, 1) or '').strip()

        ws1 = wb.sheet_by_index(0)
        if ws1.nrows > 0 and ws1.ncols > 1:
            b1 = ws1.cell_value(0, 1)
            if b1:
                info['report_number_raw'] = str(b1).strip()
                m = re.search(r'第\s*\(\s*(\d+)\s*\)\s*号', str(b1))
                if m:
                    info['report_number'] = m.group(1).strip()

        # Report date
        for r in range(9, min(14, ws1.nrows)):
            if ws1.ncols > 2:
                bv = ws1.cell_value(r, 1) if ws1.ncols > 1 else ''
                cv = ws1.cell_value(r, 2) if ws1.ncols > 2 else ''
                if bv and '报告编制日期' in str(bv) and cv:
                    info['report_date'] = str(cv).strip()
                    break

        # Sample name
        for r in range(6, min(12, ws1.nrows)):
            if ws1.ncols > 2:
                cv = ws1.cell_value(r, 2)
                if cv and ('水' in str(cv) or '【' in str(cv)):
                    info['sample_name'] = str(cv).strip()
                    break

        # Company
        for r in range(7, min(12, ws1.nrows)):
            if ws1.ncols > 2:
                cv = ws1.cell_value(r, 2)
                if cv and '公司' in str(cv):
                    info['company'] = str(cv).strip()
                    break

    return info


# ════════════════════════════════════════
# 3. 主程序
# ════════════════════════════════════════

def main():
    print("=" * 80)
    print("  报告163-180 与 原始记录(260204-1-18.xlsx) 交叉验证")
    print("=" * 80)
    print()

    # Step 1: Read original records
    print("【第一步】读取原始记录...")
    print("-" * 80)
    read_original_records()

    # Step 2: Read all report files
    print("\n\n")
    print("=" * 80)
    print("【第二步】读取报告文件 163-180...")
    print("=" * 80)

    report_files = sorted([f for f in os.listdir(BASE_DIR)
                           if f.startswith('0') and f.endswith(('.xlsx', '.xls'))
                           and not f.startswith('~')])

    reports = {}
    for fname in report_files:
        filepath = os.path.join(BASE_DIR, fname)
        prefix = re.match(r'^(\d+)', fname)
        if not prefix:
            continue
        num = int(prefix.group(1))
        if num < 163 or num > 180:
            continue

        print(f"\n--- {fname} ---")
        if fname.endswith('.xlsx'):
            info = read_xlsx_full(filepath)
        else:
            info = read_xls_full(filepath)

        reports[num] = info

        # Print key metadata
        print(f"  报告编号: {info.get('report_number', 'N/A')}")
        print(f"  样品名称: {info.get('sample_name', 'N/A')}")
        print(f"  样品编号: {info.get('sample_id', 'N/A')}")
        print(f"  样品类型: {info.get('sample_type', 'N/A')}")
        print(f"  被检单位: {info.get('company', 'N/A')}")
        print(f"  采样日期: {info.get('sampling_date', 'N/A')}")
        print(f"  收样日期: {info.get('receipt_date', 'N/A')}")
        print(f"  检测日期: {info.get('testing_date', 'N/A')}")
        print(f"  报告日期: {info.get('report_date', 'N/A')}")
        print(f"  产品标准: {info.get('product_standard', 'N/A')}")
        print(f"  检测项描述: {info.get('test_items_desc', 'N/A')}")
        print(f"  结论: {info.get('conclusion', 'N/A')[:80]}")
        print(f"  检测项目数: {len(info.get('test_items', []))}")

        # Print all test items
        for item in info.get('test_items', []):
            print(f"    {item['seq']:2d}. {item['name']:<20s} | 结果: {item['result']:<15s} | 单位: {item['unit']:<10s} | 标准: {item['standard']:<20s}")

    print(f"\n共读取 {len(reports)} 个报告文件")


if __name__ == '__main__':
    main()
